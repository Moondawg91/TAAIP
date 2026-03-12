import csv, os, json, datetime, uuid
from ..column_normalizer import normalize_col_name
from ..org_mapper import map_row_to_unit
from .. import registry
from services.api import db

IMPORTER = {
    'id': 'g2_enlistments_bn',
    'version': '1.0',
    'display_name': 'USAREC G2 - Enlistments by BN',
    'source_system': 'USAREC_G2',
    'allowed_file_types': ['xlsx','csv'],
    'detector': {
        'required_columns': ['bn','enlistments'],
        'optional_columns': ['rsid','report_date'],
        'min_required_match': 0.75,
        'min_confidence': 0.6,
        'signature_phrases': ['enlistments','battalion']
    }
}

def ensure_table():
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS g2_enlistments_bn(id TEXT PRIMARY KEY, upload_id TEXT, unit_key TEXT, rsid TEXT, bn_display TEXT, enlistments INTEGER, report_date TEXT, ingested_at TEXT)''')
        conn.commit()
    finally:
        conn.close()

def _normalize_row_map(header_row, data_row):
    # header_row: list of strings (raw headers), data_row: list of cell values
    mapping = {}
    for i, h in enumerate(header_row):
        key = normalize_col_name(h)
        mapping[key] = data_row[i] if i < len(data_row) else None
    return mapping

def load(path: str, filename: str, upload_id: str, source_system: str = None) -> dict:
    ensure_table()
    ext = os.path.splitext(filename)[1].lower()
    rows = []
    # simplistic parser: if csv, use csv.DictReader; if xlsx, fall back to import_job preview in registry
    if ext in ('.csv', '.txt'):
        with open(path, 'r', encoding='utf-8', errors='replace') as fh:
            reader = csv.reader(fh)
            all_rows = list(reader)
            # find header using registry.detect_importer logic (simple: first row with 'bn' normalized)
            header_idx = 0
            header = [c for c in all_rows[header_idx]]
            for r in all_rows[header_idx+1:]:
                mapped = _normalize_row_map(header, r)
                rows.append(mapped)
    else:
        # try to reuse registry detection preview to find header row
        det = registry.detect_importer(path, filename)
        # naive: openpyxl load to parse rows after header
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = det.get('sheet_name') or wb.sheetnames[0]
            ws = wb[sheet]
            header_row_idx = det.get('header_row_index') or 0
            rows_iter = ws.iter_rows(values_only=True)
            allrows = list(rows_iter)
            header = [str(c) if c is not None else '' for c in allrows[header_row_idx]]
            for r in allrows[header_row_idx+1:]:
                rows.append(_normalize_row_map(header, list(r)))
        except Exception:
            return {'status':'FAILED', 'inserted':0, 'updated':0, 'rejected':0, 'warnings':0, 'rejects_sample': []}

    inserted = 0
    updated = 0
    rejected = 0
    rejects_sample = []
    warnings = 0
    now = datetime.datetime.utcnow().isoformat()
    conn = db.connect()
    try:
        cur = conn.cursor()
        for idx, r in enumerate(rows, start=1):
            # mapping
            unit_map = map_row_to_unit(r)
            unit_key = unit_map.get('mapped_unit_key')
            rsid = r.get('rsid') or None
            bn_display = r.get('bn') or r.get('battalion') or None
            enlistments_raw = r.get('enlistments') or r.get('enlistment') or 0
            try:
                enlistments = int(float(str(enlistments_raw).replace(',','')))
            except Exception:
                enlistments = None
            report_date = r.get('report_date') or None
            if unit_key is None:
                rejected += 1
                if len(rejects_sample) < 25:
                    rejects_sample.append({'row': idx, 'reason':'UNMAPPED_UNIT', 'raw': r})
                # store staging_reject
                try:
                    cur.execute('INSERT INTO staging_reject(id, upload_id, dataset_id, row_number, reason_code, reason_message, raw_row_json) VALUES (?,?,?,?,?,?,?)', (uuid.uuid4().hex, upload_id, IMPORTER['id'], idx, 'UNMAPPED_UNIT', 'no unit mapping', json.dumps(r)))
                except Exception:
                    pass
                continue

            # dedupe: if same unit_key and report_date exists, update
            if report_date:
                cur.execute('SELECT id FROM g2_enlistments_bn WHERE unit_key=? AND report_date=? LIMIT 1', (unit_key, report_date))
                found = cur.fetchone()
                if found:
                    try:
                        cur.execute('UPDATE g2_enlistments_bn SET enlistments=?, rsid=?, bn_display=?, ingested_at=?, upload_id=? WHERE id=?', (enlistments, rsid, bn_display, now, upload_id, found['id'] if 'id' in found.keys() else found[0]))
                        updated += 1
                    except Exception:
                        warnings += 1
                else:
                    try:
                        cur.execute('INSERT INTO g2_enlistments_bn(id, upload_id, unit_key, rsid, bn_display, enlistments, report_date, ingested_at) VALUES (?,?,?,?,?,?,?,?)', (uuid.uuid4().hex, upload_id, unit_key, rsid, bn_display, enlistments, report_date, now))
                        inserted += 1
                    except Exception:
                        warnings += 1
            else:
                # insert with upload_id as dedupe key
                try:
                    cur.execute('INSERT INTO g2_enlistments_bn(id, upload_id, unit_key, rsid, bn_display, enlistments, report_date, ingested_at) VALUES (?,?,?,?,?,?,?,?)', (uuid.uuid4().hex, upload_id, unit_key, rsid, bn_display, enlistments, report_date, now))
                    inserted += 1
                except Exception:
                    warnings += 1
        conn.commit()
    finally:
        conn.close()

    status = 'IMPORTED' if inserted > 0 else ('PARTIAL' if rejected > 0 else 'FAILED')
    return {'status': status, 'inserted': inserted, 'updated': updated, 'rejected': rejected, 'warnings': warnings, 'rejects_sample': rejects_sample}
