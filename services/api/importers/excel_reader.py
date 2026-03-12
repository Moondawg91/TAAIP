import csv, io
try:
    import openpyxl
except Exception:
    openpyxl = None

def preview_xlsx(path, max_rows=60, max_cols=60):
    if openpyxl is None:
        raise RuntimeError('openpyxl missing')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    previews = {}
    for sheet in wb.sheetnames[:3]:
        ws = wb[sheet]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append([cell for cell in r[:max_cols]])
            if i+1 >= max_rows: break
        previews[sheet] = rows
    return previews

def preview_csv(path, max_rows=120):
    out = []
    with open(path, 'r', encoding='utf-8', errors='replace') as fh:
        # basic delimiter sniff
        sample = fh.read(8192)
        fh.seek(0)
        dialect = None
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        for i, r in enumerate(reader):
            out.append(r)
            if i+1 >= max_rows: break
    return {'sheet0': out}
