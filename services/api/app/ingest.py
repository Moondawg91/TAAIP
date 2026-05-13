import os
import json
import pandas as pd
from .database import SessionLocal
from . import models_ingest, models
from sqlalchemy.orm import Session
import re
from dateutil import parser as date_parser
from . import ingest_registry
from .db import get_db_conn
from sqlalchemy import text


def read_file_to_df(path: str):
    # try excel, then csv, then json
    try:
        if path.lower().endswith('.xlsx') or path.lower().endswith('.xls'):
            return pd.read_excel(path, engine='openpyxl')
        if path.lower().endswith('.csv'):
            return pd.read_csv(path)
        if path.lower().endswith('.json'):
            return pd.read_json(path)
    except Exception:
        raise
    raise ValueError('Unsupported file type')


def apply_recipe_to_df(df: pd.DataFrame, steps: list) -> (pd.DataFrame, dict):
    report = {"steps": [], "rows_before": len(df), "rows_after": None}
    cur = df.copy()
    for step in steps:
        stype = step.get('type')
        if stype == 'cast':
            col = step['column']
            to_type = step['to']
            try:
                if to_type == 'int':
                    cur[col] = pd.to_numeric(cur[col], errors='coerce').astype('Int64')
                elif to_type == 'float':
                    cur[col] = pd.to_numeric(cur[col], errors='coerce')
                elif to_type == 'str':
                    cur[col] = cur[col].astype(str)
                report['steps'].append({'step': step, 'status': 'ok'})
            except Exception as e:
                report['steps'].append({'step': step, 'status': 'error', 'error': str(e)})
        elif stype == 'filter':
            expr = step['expr']
            try:
                cur = cur.query(expr)
                report['steps'].append({'step': step, 'status': 'ok'})
            except Exception as e:
                report['steps'].append({'step': step, 'status': 'error', 'error': str(e)})
        elif stype == 'dedupe':
            cols = step.get('columns')
            cur = cur.drop_duplicates(subset=cols)
            report['steps'].append({'step': step, 'status': 'ok'})
        elif stype == 'map':
            col = step['column']
            mapping = step.get('mapping', {})
            cur[col] = cur[col].map(mapping).fillna(cur[col])
            report['steps'].append({'step': step, 'status': 'ok'})
        else:
            report['steps'].append({'step': step, 'status': 'skipped', 'reason': 'unknown step type'})

    report['rows_after'] = len(cur)
    return cur, report


def save_ingested_file(db: Session, filename: str, source: str, uploaded_by: str):
    f = models_ingest.IngestedFile(filename=filename, source=source, uploaded_by=uploaded_by)
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


def run_ingest(path: str, recipe: dict = None, uploaded_by: str = None):
    db = SessionLocal()
    try:
        ing = save_ingested_file(db, os.path.basename(path), path, uploaded_by)
        steps = recipe.get('steps') if recipe else []
        df = read_file_to_df(path)
        out_df, report = apply_recipe_to_df(df, steps)
        run = models_ingest.IngestRun(file_id=ing.id, recipe_id=recipe.get('id') if recipe else None, status='completed', report=report)
        db.add(run)
        db.commit()
        db.refresh(run)
        return {'file': ing.id, 'run': run.id, 'report': report, 'preview': out_df.head(20).to_dict(orient='records')}
    finally:
        db.close()


def profile_file(path: str) -> dict:
    """Return a lightweight profile: sheets, columns per sheet, and sample rows."""
    prof = {"sheets": [], "columns": [], "sample": []}
    try:
        if path.lower().endswith('.xlsx') or path.lower().endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                # read header row (first non-empty)
                header = None
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    if row and any([c is not None for c in row]):
                        header = [str(c).strip() if c is not None else '' for c in row]
                        break
                prof['sheets'].append(name)
                if header:
                    prof['columns'].append({"sheet": name, "columns": header})
                # sample first few rows
                sample = []
                rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
                for r in rows:
                    sample.append([c for c in r])
                prof['sample'].append({"sheet": name, "rows": sample})
        elif path.lower().endswith('.csv'):
            import csv
            with open(path, newline='') as fh:
                rdr = csv.reader(fh)
                hdr = next(rdr, [])
                prof['sheets'].append('csv')
                prof['columns'].append({"sheet": 'csv', "columns": [str(h).strip() for h in hdr]})
                sample = []
                for i, row in enumerate(rdr):
                    if i >= 5:
                        break
                    sample.append(row)
                prof['sample'].append({"sheet": 'csv', "rows": sample})
    except Exception:
        pass
    return prof


def detect_importer(path: str):
    prof = profile_file(path)
    # flatten columns
    cols = []
    for c in prof.get('columns', []):
        cols.extend([str(x).upper() for x in c.get('columns', []) if x])
    sheets = [s.upper() for s in prof.get('sheets', [])]
    for spec in ingest_registry.IMPORTERS:
        fp = spec.get('fingerprint', {})
        # check sheet hints
        hints = [h.upper() for h in fp.get('sheetNameHints', [])] if fp.get('sheetNameHints') else []
        # For CSV inputs (sheets == ['CSV']), allow matching based on columns
        # even if the importer defines sheet name hints. This keeps CSV
        # detection permissive for simple exports while keeping hint checks
        # for Excel/complex workbooks.
        if hints and 'CSV' not in sheets:
            # require any hint present in sheet names
            if not any(any(h in s for s in sheets) for h in hints):
                continue
        # requiredColumnsAnyOf: list of lists, any one list satisfied
        sets = fp.get('requiredColumnsAnyOf', [])
        matched = False
        for sset in sets:
            if all(any(sc.upper() == c or sc.upper() in c for c in cols) for sc in sset):
                matched = True
                break
        if matched:
            return spec['id']
    return None


def _normalize_value(val, dtype):
    if val is None:
        return None
    try:
        s = str(val).strip()
    except Exception:
        s = val
    if dtype == 'int':
        try:
            return int(str(s).replace(',', ''))
        except Exception:
            return None
    if dtype == 'float' or dtype == 'percent':
        try:
            vv = str(s).replace(',', '').replace('%', '')
            v = float(vv)
            if dtype == 'percent':
                return v / 100.0
            return v
        except Exception:
            return None
    return s


def run_import(path: str, ingest_run_id: int = None, importer_id: str = None, db: Session = None, uploaded_by: str = None) -> dict:
    """Run import for a given file path and importer spec. Returns summary dict."""
    result = {"status": "completed", "row_count_in": 0, "row_count_loaded": 0, "errors": []}
    conn = None
    cur = None

    def _fetchone(sql: str, params: dict = None):
        if db is not None:
            rr = db.execute(text(sql), params or {}).fetchone()
            try:
                return dict(rr) if rr is not None else None
            except Exception:
                return rr
        c = cur.execute(sql, params or {})
        rr = c.fetchone()
        if rr is None:
            return None
        try:
            return dict(rr)
        except Exception:
            return rr

    def _execute(sql: str, params: dict = None):
        if db is not None:
            return db.execute(text(sql), params or {})
        return cur.execute(sql, params or {})

    def _commit():
        if db is not None:
            db.commit()
        else:
            conn.commit()

    if db is None:
        conn = get_db_conn()
        cur = conn.cursor()

    def _table_columns(table_name: str):
        try:
            if db is not None:
                rows = db.execute(text(f"PRAGMA table_info('{table_name}')")).fetchall()
            else:
                rows = cur.execute(f"PRAGMA table_info('{table_name}')").fetchall()
            out = []
            for rr in rows:
                try:
                    out.append(rr[1])
                except Exception:
                    try:
                        out.append(rr.get('name'))
                    except Exception:
                        pass
            return [c for c in out if c]
        except Exception:
            return []
    try:
        df = read_file_to_df(path)
    except Exception as e:
        raise
    # if importer specified, load spec
    spec = None
    if importer_id:
        spec = ingest_registry.get_importer(importer_id)
    else:
        importer_id = detect_importer(path)
        spec = ingest_registry.get_importer(importer_id) if importer_id else None
    if not spec:
        # unknown dataset: persist to stg_raw_dataset rows
        rows = df.to_dict(orient='records') if hasattr(df, 'to_dict') else []
        # Prefer using the shared SQLAlchemy session when available to
        # avoid mixing independent sqlite3 connections which can hold
        # locks during concurrent test runs.
        if db:
            stmt = text("INSERT INTO stg_raw_dataset (ingest_run_id, row_number, row_json) VALUES (:ingest_run_id, :row_number, :row_json)")
            for i, r in enumerate(rows, start=1):
                db.execute(stmt, { 'ingest_run_id': ingest_run_id, 'row_number': i, 'row_json': json.dumps(r) })
            try:
                db.commit()
            except Exception:
                pass
        else:
            for i, r in enumerate(rows, start=1):
                cur.execute("INSERT INTO stg_raw_dataset (ingest_run_id, row_number, row_json) VALUES (?, ?, ?)", (ingest_run_id, i, json.dumps(r)))
            conn.commit()
        result['status'] = 'staged_unknown'
        result['row_count_in'] = len(rows)
        return result

    # map headers to canonical columns
    hdr_map = {}
    # build mapping from aliases to canonical
    for col in spec.get('columns', []):
        for a in col.get('aliases', []):
            hdr_map[a.upper()] = col['canonical']
        # also include canonical itself
        hdr_map[col['canonical'].upper()] = col['canonical']

    # normalize df columns to canonical keys
    orig_cols = list(df.columns)
    mapped_cols = {}
    for c in orig_cols:
        key = str(c).upper()
        if key in hdr_map:
            mapped_cols[c] = hdr_map[key]
        else:
            mapped_cols[c] = str(c)
    df = df.rename(columns=mapped_cols)
    rows = df.to_dict(orient='records')
    result['row_count_in'] = len(rows)

    # apply transforms per spec
    out_rows = []
    for idx, r in enumerate(rows, start=1):
        try:
            ctx = dict(r)
            # apply each transform
            for t in spec.get('transforms', []):
                fn = t.get('fn')
                args = t.get('args', {})
                if fn == 'normalize_unit':
                    # attempt rsid direct then lookup by name
                    to_field = args.get('to', 'unit_rsid')
                    from_fields = args.get('from') if isinstance(args.get('from'), list) else [args.get('from')]
                    found = None
                    for f in from_fields:
                        v = ctx.get(f)
                        if not v:
                            continue
                        # exact rsid
                        row = _fetchone("SELECT rsid FROM org_unit WHERE upper(rsid)=:v LIMIT 1", {"v": str(v).upper()})
                        if row:
                            found = row['rsid'] if isinstance(row, dict) else row[0]
                            break
                        # name/display_name contains (schema-compatible)
                        org_cols = set(_table_columns('org_unit'))
                        if 'display_name' in org_cols:
                            row = _fetchone("SELECT rsid FROM org_unit WHERE upper(display_name) LIKE :v LIMIT 1", {"v": f"%{str(v).upper()}%"})
                            if row:
                                found = row['rsid'] if isinstance(row, dict) else row[0]
                                break
                        if 'name' in org_cols:
                            row = _fetchone("SELECT rsid FROM org_unit WHERE upper(name) LIKE :v LIMIT 1", {"v": f"%{str(v).upper()}%"})
                            if row:
                                found = row['rsid'] if isinstance(row, dict) else row[0]
                                break
                    ctx[to_field] = found
                elif fn == 'normalize_dates':
                    from_field = args.get('from')
                    to_fields = args.get('to', [])
                    val = ctx.get(from_field)
                    if val:
                        try:
                            dt = date_parser.parse(str(val), default=None)
                            # for period strings that are months, choose start/end approximations
                            start = dt.isoformat()
                            end = dt.isoformat()
                            if len(to_fields) >= 1:
                                ctx[to_fields[0]] = start
                            if len(to_fields) >= 2:
                                ctx[to_fields[1]] = end
                        except Exception:
                            pass
                elif fn == 'derive_grain':
                    ctx['grain'] = args.get('grain')
                elif fn == 'map_metrics_wide_to_long':
                    mapping = args.get('map', [])
                    # produce multiple metric rows per source row
                    for m in mapping:
                        from_col = m.get('from')
                        metric = m.get('metric')
                        val = ctx.get(from_col)
                        new = dict(ctx)
                        new['metric_name'] = metric
                        new['metric_value'] = _normalize_value(val, 'float')
                        out_rows.append(new)
                    # skip adding ctx below
                    ctx = None
                elif fn == 'zip_to_unit_lookup':
                    zip_field = args.get('zipField')
                    rsid_field = args.get('rsidField')
                    to = args.get('to', 'unit_rsid')
                    z = ctx.get(zip_field)
                    rs = ctx.get(rsid_field)
                    found = None
                    if rs:
                        row = _fetchone("SELECT rsid FROM org_unit WHERE upper(rsid)=:v LIMIT 1", {"v": str(rs).upper()})
                        if row:
                            found = row['rsid'] if isinstance(row, dict) else row[0]
                    if not found and z:
                        # placeholder lookup by zip -> org_unit (best-effort)
                        row = _fetchone("SELECT rsid FROM org_unit WHERE location_zip = :z LIMIT 1", {"z": str(z)})
                        if row:
                            found = row['rsid'] if isinstance(row, dict) else row[0]
                    ctx[to] = found
                elif fn == 'cbsa_lookup':
                    # upsert cbsa dim
                    code = ctx.get(args.get('codeField'))
                    name = ctx.get(args.get('nameField'))
                    if code:
                        if _table_columns('dim_market_cbsa'):
                            _execute(
                                "INSERT OR REPLACE INTO dim_market_cbsa (cbsa_code, cbsa_name, state, urbanicity_pct, updated_at) VALUES (:code, :name, :st, :u, datetime('now'))",
                                {"code": code, "name": name, "st": None, "u": None},
                            )
                            _commit()
            if ctx is not None:
                out_rows.append(ctx)
        except Exception as e:
            result['errors'].append({"row": idx, "error": str(e)})

    # load out_rows into target table(s)
    loaded = 0
    target = spec.get('target', {})
    ttable = target.get('table')
    mode = target.get('mode')
    pkeys = target.get('primaryKey', [])
    table_cols = set(_table_columns(ttable))

    def _shape_row_for_table(row: dict) -> dict:
        shaped = {k: v for k, v in row.items() if k in table_cols}

        # Compatibility mappings for legacy canonical schemas.
        if ttable == 'fact_enlistments':
            if 'unit_rsid' in table_cols and not shaped.get('unit_rsid'):
                shaped['unit_rsid'] = row.get('unit_rsid') or row.get('rsid')
            if 'echelon' in table_cols and not shaped.get('echelon'):
                shaped['echelon'] = row.get('grain') or row.get('echelon')
            if 'period_date' in table_cols and not shaped.get('period_date'):
                shaped['period_date'] = row.get('period_start') or row.get('period') or row.get('period_end')
            if 'contracts' in table_cols and shaped.get('contracts') is None:
                if str(row.get('metric_name') or '').lower() in {'enlistments', 'contracts'}:
                    shaped['contracts'] = row.get('metric_value')
                elif row.get('enlistments') is not None:
                    shaped['contracts'] = row.get('enlistments')
                elif row.get('value') is not None:
                    shaped['contracts'] = row.get('value')

        if ttable == 'fact_zip_potential':
            if 'zip' in table_cols and not shaped.get('zip'):
                shaped['zip'] = row.get('zip') or row.get('zip5')
            if 'category' in table_cols and not shaped.get('category'):
                shaped['category'] = row.get('category') or row.get('market_category')
            if 'metric_name' in table_cols and not shaped.get('metric_name'):
                shaped['metric_name'] = row.get('metric_name') or 'value'
            if 'metric_value' in table_cols and shaped.get('metric_value') is None:
                shaped['metric_value'] = row.get('metric_value') if row.get('metric_value') is not None else row.get('value')

        if 'source_system' in table_cols and not shaped.get('source_system'):
            shaped['source_system'] = spec.get('sourceSystem') or spec.get('dataset_key') or importer_id
        if 'dataset_key' in table_cols and not shaped.get('dataset_key'):
            shaped['dataset_key'] = spec.get('dataset_key') or importer_id
        if 'ingest_run_id' in table_cols and not shaped.get('ingest_run_id'):
            shaped['ingest_run_id'] = ingest_run_id

        # remove explicit id to avoid accidental PK collisions on AUTOINCREMENT tables
        shaped.pop('id', None)
        return shaped

    for r in out_rows:
        try:
            insert_row = _shape_row_for_table(r)
            if not insert_row:
                continue

            # if upsert mode and primary keys provided, delete existing
            if mode == 'upsert' and pkeys:
                usable_keys = [k for k in pkeys if k in table_cols and insert_row.get(k) is not None]
                if usable_keys:
                    where = ' AND '.join([f"{k} = :k_{k}" for k in usable_keys])
                    params = {f"k_{k}": insert_row.get(k) for k in usable_keys}
                    _execute(f"DELETE FROM {ttable} WHERE {where}", params)
            # build insert columns from r keys
            cols = ','.join(insert_row.keys())
            val_names = [f"v_{i}" for i, _ in enumerate(insert_row.keys())]
            vals = ','.join([f":{vn}" for vn in val_names])
            params = {vn: list(insert_row.values())[i] for i, vn in enumerate(val_names)}
            _execute(f"INSERT INTO {ttable} ({cols}) VALUES ({vals})", params)
            loaded += 1
        except Exception as e:
            result['errors'].append({"row": r, "error": str(e)})
    _commit()
    result['row_count_loaded'] = loaded
    if db is None and conn is not None:
        try:
            conn.close()
        except Exception:
            pass
    return result
