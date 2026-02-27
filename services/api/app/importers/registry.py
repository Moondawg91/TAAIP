"""Importer registry and lightweight parsers for Data Hub.

This module defines a small set of importer specs and provides
`detect_importer()` and `run_import()` used by the router.
"""
from typing import List, Dict, Any
import json
import io
import csv
import re
import difflib
from openpyxl import load_workbook
from .. import db as _db
from .unit_mapping import map_unit_rsid
from datetime import datetime


def _now_iso():
    return datetime.utcnow().isoformat()


# Lightweight spec definitions. Column names are lowercase-normalized.
SPECS = [
    {
        'source_system': 'EMM',
        'dataset_key': 'emm_events',
        'friendly_name': 'EMM Portal - Events',
        'file_name_patterns': ['emm', 'events', 'mac'],
        'sheet_name_patterns': ['events', 'mac', 'sheet1'],
        'header_signatures': [
            {'event_name', 'start_date', 'end_date'},
            {'event', 'start', 'end'},
            {'event_name', 'location', 'city'}
        ]
    },
    {
        'source_system': 'EMM',
        'dataset_key': 'emm_macs',
        'friendly_name': 'EMM Portal - MACs',
        'file_name_patterns': ['emm','macs'],
        'sheet_name_patterns': ['macs','contacts'],
        'header_signatures': [
            {'mac_id', 'mac_name', 'mac_type'},
        ]
    },
    {
        'source_system': 'ALRL',
        'dataset_key': 'alrl_schools',
        'friendly_name': 'ALRL - Schools',
        'file_name_patterns': ['alrl', 'school','schools'],
        'header_signatures': [
            {'school_name', 'city', 'state'},
            {'school', 'district', 'city'}
        ]
    },
    {
        'source_system': 'USAREC_G2',
        'dataset_key': 'g2_market_metric',
        'friendly_name': 'USAREC G2 - Metrics (various shapes)',
        'file_name_patterns': ['g2', 'enlist', 'productivity', 'zip','sama','cbsa'],
        'header_signatures': [
            {'cbsa', 'enlistments'},
            {'zip', 'zip_category', 'potential'},
            {'bde', 'bn', 'enlistments'},
            {'station', 'enlistments'},
            {'urbanicity', 'pct'}
        ]
    },
    {
        'source_system': 'AIE',
        'dataset_key': 'aie_leads_v1',
        'friendly_name': 'AIE - Leads (stub)',
        'file_name_patterns': ['aie', 'lead', 'person'],
        'header_signatures': [
            {'person', 'lead', 'source'},
        ]
    }
]


def _normalize_headers(headers: List[str]) -> List[str]:
    return [ (h or '').strip().lower() for h in headers ]


def _peek_csv_columns(b):
    try:
        s = b.decode('utf-8', errors='ignore')
        rdr = csv.reader(io.StringIO(s))
        for row in rdr:
            if row and any(cell.strip() for cell in row):
                return _normalize_headers(row)
    except Exception:
        pass
    return []


def _peek_xlsx_columns(b):
    try:
        wb = load_workbook(filename=io.BytesIO(b), read_only=True)
        results = []
        for name in wb.sheetnames:
            sheet = wb[name]
            # find first non-empty row within first 8 rows
            for r_idx, r in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
                if r and any((c is not None and str(c).strip() != '') for c in r):
                    results.append({'sheet': name, 'header_row': r_idx-1, 'columns': _normalize_headers([str(c) if c is not None else '' for c in r])})
                    break
        return results
    except Exception:
        pass
    return []


def list_specs():
    return [ {k: spec[k] for k in ('source_system','dataset_key','friendly_name')} for spec in SPECS ]


def detect_importer(file_bytes: bytes, filename: str) -> Dict[str,Any]:
    fname = (filename or '').lower()
    csv_cols = _peek_csv_columns(file_bytes)
    xlsx_peeks = _peek_xlsx_columns(file_bytes)
    best = None
    best_score = 0
    best_details = {}

    # Helper to score against a column list
    def score_against(cols_list, spec):
        score = 0
        cols_set = set(cols_list or [])
        # filename patterns
        for p in spec.get('file_name_patterns', []):
            if p in fname:
                score += 3
        # sheet name patterns (if provided in spec and available)
        for sp in spec.get('sheet_name_patterns', []):
            if xlsx_peeks and any(sp in n.get('sheet','').lower() for n in xlsx_peeks):
                score += 4
        # header signatures
        for sig in spec.get('header_signatures', []):
            if cols_set and set(sig).issubset(cols_set):
                score += 20
            else:
                # partial matches get smaller credit
                score += len(set(sig).intersection(cols_set)) * 2
        return score

    # Score CSV candidate if any
    if csv_cols:
        for spec in SPECS:
            s = score_against(csv_cols, spec)
            if s > best_score:
                best_score = s
                best = spec
                best_details = {'detected_columns': csv_cols, 'sheet': None, 'header_row': 0}

    # Score XLSX sheet-level candidates
    for peek in xlsx_peeks:
        cols = peek.get('columns')
        sheet = peek.get('sheet')
        header_row = peek.get('header_row', 0)
        for spec in SPECS:
            s = score_against(cols, spec)
            # give small bonus if sheet name explicitly matches patterns
            for sp in spec.get('sheet_name_patterns', []):
                if sp in (sheet or '').lower():
                    s += 3
            if s > best_score:
                best_score = s
                best = spec
                best_details = {'detected_columns': cols, 'sheet': sheet, 'header_row': header_row}

    detection = {
        'source_system': best['source_system'] if best else 'UNKNOWN',
        'dataset_key': best['dataset_key'] if best else 'unknown',
        'confidence': float(best_score),
        'detected_columns': best_details.get('detected_columns') if best_details else (csv_cols or []),
        'filename': filename,
        'sheet': best_details.get('sheet'),
        'header_row': best_details.get('header_row', 0)
    }
    return detection


def run_import(detection: dict, file_bytes: bytes, import_run_id: int, dry_run: bool = True) -> dict:
    """Parse and optionally commit normalized rows.

    Returns a dict with keys: rows_in, rows_inserted, rows_rejected, warnings, errors, preview
    """
    cols = detection.get('detected_columns', [])
    rows_in = 0
    rows_inserted = 0
    rows_rejected = 0
    warnings = []
    errors = []
    preview = []

    # Simple parser: support CSV and XLSX (first sheet) and produce row dicts
    parsed_rows = []
    # try CSV
    try:
        s = file_bytes.decode('utf-8')
        rdr = csv.DictReader(io.StringIO(s))
        for r in rdr:
            parsed_rows.append({k.strip(): (v if v is not None else '') for k,v in r.items()})
    except Exception:
        # try xlsx
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
            sheet = wb[wb.sheetnames[0]]
            header = None
            for r in sheet.iter_rows(values_only=True):
                if header is None:
                    header = [ (c or '').strip() for c in r ]
                    continue
                if not any(c is not None and str(c).strip() != '' for c in r):
                    continue
                row = {}
                for i, cell in enumerate(r):
                    key = header[i] if i < len(header) else f'col_{i}'
                    row[key] = cell
                parsed_rows.append(row)
        except Exception as e:
            errors.append(str(e))

    rows_in = len(parsed_rows)
    # store raw staging rows
    conn = _db.connect()
    cur = conn.cursor()
    for i, row in enumerate(parsed_rows, start=1):
        try:
            cur.execute("INSERT INTO stg_raw_dataset (ingest_run_id, row_number, row_json) VALUES (?, ?, ?)", (import_run_id, i, json.dumps(row, default=str)))
        except Exception:
            pass
    conn.commit()

    # Preview first 10 rows
    for r in parsed_rows[:10]:
        preview.append(r)

    # If dry-run, just return validation info
    if dry_run:
        return {
            'rows_in': rows_in,
            'rows_inserted': 0,
            'rows_rejected': 0,
            'warnings': warnings,
            'errors': errors,
            'preview': preview
        }

    # Commit normalized rows based on dataset_key
    ds = detection.get('dataset_key', '')
    if ds == 'emm_events' or detection.get('source_system') == 'EMM':
        for i, row in enumerate(parsed_rows, start=1):
            # map common columns
            ev_name = row.get('event_name') or row.get('event') or row.get('name')
            start = row.get('start_date') or row.get('start')
            end = row.get('end_date') or row.get('end')
            city = row.get('city')
            state = row.get('state')
            zipc = row.get('zip') or row.get('zipcode')
            unit_rsid, confidence, reason = map_unit_rsid(row)
            if confidence and confidence < 1.0:
                # attach a mapping warning
                if confidence > 0:
                    warnings.append({'row': i, 'warning': f'unit_mapped_low_confidence:{reason}:{confidence}'})
                else:
                    warnings.append({'row': i, 'warning': f'unit_unmapped:{reason}'})
            try:
                cur.execute(
                    "INSERT INTO emm_event (event_id,event_name,start_date,end_date,city,state,zip,unit_rsid,source_import_run_id,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (None, ev_name, start, end, city, state, zipc, unit_rsid, import_run_id, _now_iso())
                )
                rows_inserted += 1
            except Exception as e:
                rows_rejected += 1
                cur.execute("INSERT INTO import_row_error (import_run_id, row_number, severity, message, raw_row_json) VALUES (?,?,?,?,?)", (import_run_id, i, 'ERROR', str(e), json.dumps(row, default=str)))
    elif ds == 'alrl_schools' or detection.get('source_system') == 'ALRL':
        for i, row in enumerate(parsed_rows, start=1):
            try:
                cur.execute(
                    "INSERT INTO alrl_school (school_id,school_name,city,state,zip,source_import_run_id) VALUES (?,?,?,?,?,?)",
                    (row.get('school_id') or None, row.get('school_name') or row.get('school') or row.get('name'), row.get('city'), row.get('state'), row.get('zip'), import_run_id)
                )
                rows_inserted += 1
            except Exception as e:
                rows_rejected += 1
                cur.execute("INSERT INTO import_row_error (import_run_id, row_number, severity, message, raw_row_json) VALUES (?,?,?,?,?)", (import_run_id, i, 'ERROR', str(e), json.dumps(row, default=str)))
    elif ds == 'g2_market_metric' or detection.get('source_system') == 'USAREC_G2':
        for i, row in enumerate(parsed_rows, start=1):
            # pick a likely metric column
            metric = None
            value = None
            for k in row.keys():
                kl = str(k).strip().lower()
                if 'enlist' in kl or 'enlistments' in kl or 'potential' in kl or 'metric' in kl:
                    metric = kl
                    try:
                        value = float(row[k])
                    except Exception:
                        value = str(row[k])
                    break
            if metric is None:
                metric = 'unknown'
                value = json.dumps(row)
            # try mapping unit
            unit_rsid, conf, reason = map_unit_rsid(row)
            try:
                if isinstance(value, float):
                    cur.execute("INSERT INTO g2_market_metric (metric_key,value_real,as_of_date,unit_rsid,source_import_run_id) VALUES (?,?,?,?,?)", (metric, value, None, unit_rsid, import_run_id))
                else:
                    cur.execute("INSERT INTO g2_market_metric (metric_key,value_text,as_of_date,unit_rsid,source_import_run_id) VALUES (?,?,?,?,?)", (metric, str(value), None, unit_rsid, import_run_id))
                rows_inserted += 1
            except Exception as e:
                rows_rejected += 1
                cur.execute("INSERT INTO import_row_error (import_run_id, row_number, severity, message, raw_row_json) VALUES (?,?,?,?,?)", (import_run_id, i, 'ERROR', str(e), json.dumps(row, default=str)))
    elif ds == 'aie_leads_stub' or detection.get('source_system') == 'AIE':
        for i, row in enumerate(parsed_rows, start=1):
            try:
                # flexible mapping into canonical lead_journey_fact
                lead_id = row.get('lead_id') or row.get('id') or row.get('person_key') or None
                person_key = row.get('person_key') or row.get('person') or None
                unit_rsid, conf, reason = map_unit_rsid(row)
                source_type = (row.get('source_type') or row.get('source') or 'aie').strip()
                source_detail = row.get('campaign') or row.get('channel') or row.get('source_detail') or None
                lead_created = row.get('created_at') or row.get('lead_created_dt') or row.get('lead_created') or None
                contact_made = row.get('contact_made_dt') or row.get('contacted_at') or None
                appointment = row.get('appointment_dt') or row.get('appt_date') or None
                applicant = row.get('applicant_dt') or row.get('applicant_date') or None
                contract_dt = row.get('contract_dt') or row.get('contract_date') or None
                contract_flag = 1 if (row.get('contract_flag') or row.get('contract') or '').lower() in ('1','true','yes') else 0
                contract_type = row.get('contract_type') or None
                mos = row.get('mos') or None
                afqt_tier = row.get('afqt_tier') or None
                now = _now_iso()
                cur.execute(
                    "INSERT OR REPLACE INTO lead_journey_fact (lead_id, person_key, unit_rsid, source_type, source_detail, event_id, mac_id, lead_created_dt, contact_made_dt, appointment_dt, applicant_dt, contract_dt, contract_flag, contract_type, mos, afqt_tier, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (lead_id, person_key, unit_rsid, source_type, source_detail, None, None, lead_created, contact_made, appointment, applicant, contract_dt, contract_flag, contract_type, mos, afqt_tier, now, now)
                )
                rows_inserted += 1
                if conf and conf < 1.0:
                    if conf > 0:
                        warnings.append({'row': i, 'warning': f'unit_mapped_low_confidence:{reason}:{conf}'})
                    else:
                        warnings.append({'row': i, 'warning': f'unit_unmapped:{reason}'})
            except Exception as e:
                rows_rejected += 1
                cur.execute("INSERT INTO import_row_error (import_run_id, row_number, severity, message, raw_row_json) VALUES (?,?,?,?,?)", (import_run_id, i, 'ERROR', str(e), json.dumps(row, default=str)))
    else:
        # fallback: write rows into stg_raw_dataset_profile or stg_raw_dataset
        for i, row in enumerate(parsed_rows, start=1):
            try:
                cur.execute("INSERT INTO stg_raw_dataset (ingest_run_id, row_number, row_json) VALUES (?,?,?)", (import_run_id, i, json.dumps(row, default=str)))
                rows_inserted += 1
            except Exception as e:
                rows_rejected += 1
                cur.execute("INSERT INTO import_row_error (import_run_id, row_number, severity, message, raw_row_json) VALUES (?,?,?,?,?)", (import_run_id, i, 'ERROR', str(e), json.dumps(row, default=str)))

    conn.commit()
    return {
        'rows_in': rows_in,
        'rows_inserted': rows_inserted,
        'rows_rejected': rows_rejected,
        'warnings': warnings,
        'errors': errors,
        'preview': preview,
        'success': (len(errors) == 0)
    }
