from fastapi import APIRouter, UploadFile, File, HTTPException, Form, Depends, Body
from typing import Dict, List, Optional, Any
from .. import db
from .rbac import require_scope, get_current_user
import os, hashlib, json, csv, io, datetime, re, uuid, pathlib
import sqlite3

try:
    import openpyxl
except Exception:
    openpyxl = None

router = APIRouter()

# small alias map to translate common mapping names to DB column names
FIELD_ALIASES = {
    'date': 'date_key',
    'date_key': 'date_key',
    'org_unit': 'org_unit_id',
    'org_unit_id': 'org_unit_id',
    'count': 'count_value',
    'count_value': 'count_value'
}

def now_iso():
    return datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')


def audit(conn, who: str, action: str, entity: str, entity_id: int = None, meta: dict = None):
    try:
        cur = conn.cursor()
        cur.execute('INSERT INTO audit_log(who, action, entity, entity_id, meta_json, created_at) VALUES (?,?,?,?,?,?)', (
            who or 'system', action, entity, entity_id, json.dumps(meta or {}), now_iso()
        ))
        conn.commit()
    except Exception:
        pass


def _update_import_job_status_fallback(cur, job_id, status, notes=None):
    """Try to update import_job including notes; fall back if legacy schema lacks notes."""
    try:
        if notes is not None:
            cur.execute('UPDATE import_job SET status=?, notes=?, updated_at=? WHERE id=?', (status, notes, now_iso(), job_id))
        else:
            cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', (status, now_iso(), job_id))
    except sqlite3.OperationalError:
        try:
            cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', (status, now_iso(), job_id))
        except Exception:
            pass


@router.post('/import/upload')
async def upload_file(file: UploadFile = File(...), uploaded_by: Optional[str] = None, target_domain: Optional[str] = 'generic', dataset: Optional[str] = Form(None), allowed_orgs: Optional[list] = Depends(require_scope('STATION'))) -> Dict:
    # create job row first so we can store file under a safe per-job folder
    conn = db.connect()
    try:
        cur = conn.cursor()
        created_at = now_iso()
        cur.execute("INSERT INTO import_job(filename_original, file_type, file_size_bytes, sha256_hash, uploaded_by_user_id, uploaded_at, status, target_domain, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (file.filename, os.path.splitext(file.filename)[1].lstrip('.').lower(), 0, None, uploaded_by, created_at, 'uploaded', target_domain, created_at, created_at))
        jid = cur.lastrowid
        conn.commit()
        try:
            audit(conn, uploaded_by or 'uploader', 'import.upload', 'import_job', jid, {'filename': file.filename, 'target_domain': target_domain})
        except Exception:
            pass
    finally:
        conn.close()

    # If client provided a `dataset` form field (legacy client), prefer it as target_domain
    try:
        if dataset:
            target_domain = dataset
    except Exception:
        pass

    # store file in per-job folder
    base_dir = os.getenv('TAAIP_UPLOAD_DIR', 'services/api/.data/imports')
    job_dir = os.path.join(base_dir, f"job_{jid}")
    os.makedirs(job_dir, exist_ok=True)
    contents = await file.read()

    # Hardening: reject imports that contain simulation/demo markers unless explicitly allowed
    try:
        if os.getenv('ALLOW_SIMULATION_IMPORTS') != '1':
            sim_pat = re.compile(r"\bSIM_|\bsim-|\bdemo-|\bdemo_", re.IGNORECASE)
            try:
                s = contents.decode('utf-8', errors='ignore')
            except Exception:
                s = ''
            if sim_pat.search(s):
                raise HTTPException(status_code=400, detail="Import rejected: contains simulation/demo markers. Set ALLOW_SIMULATION_IMPORTS=1 to override.")
    except HTTPException:
        raise
    except Exception:
        # conservative: if content inspection fails, reject the upload
        raise HTTPException(status_code=400, detail="Import validation failed; refused to process file")
    sha = hashlib.sha256(contents).hexdigest()
    fname = f"original{os.path.splitext(file.filename)[1].lower()}"
    path = os.path.join(job_dir, fname)
    with open(path, 'wb') as fh:
        fh.write(contents)

    # update job with sha and file size
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('UPDATE import_job SET sha256_hash=?, file_size_bytes=?, file_type=?, updated_at=? WHERE id=?', (sha, len(contents), os.path.splitext(file.filename)[1].lstrip('.').lower(), now_iso(), jid))
        # also create a v3 import_job id for provenance tracking
        import_job_uuid = uuid.uuid4().hex
        try:
            cur.execute('INSERT OR REPLACE INTO import_job_v3(id, created_at, created_by, dataset_key, source_system, filename, file_sha256, status, row_count, error_count, notes, scope_org_unit_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', (
                import_job_uuid, now_iso(), uploaded_by, target_domain or 'generic', None, file.filename, sha, 'uploaded', 0, 0, None, None
            ))
        except Exception:
            pass
        # record import_file metadata
        try:
            # Prefer modern schema (sha256, original_filename, stored_path, content_type, byte_size, uploaded_at)
            cur.execute('INSERT OR REPLACE INTO import_file(sha256, original_filename, stored_path, content_type, byte_size, uploaded_at) VALUES (?,?,?,?,?,?)', (
                sha, file.filename, path, file.content_type or os.path.splitext(file.filename)[1].lstrip('.').lower(), len(contents), now_iso()
            ))
        except Exception:
            try:
                # Fallback legacy schema (id, import_job_id, stored_path, content_type, size_bytes, uploaded_at)
                file_id = uuid.uuid4().hex
                cur.execute('INSERT OR REPLACE INTO import_file(id, import_job_id, stored_path, content_type, size_bytes, uploaded_at) VALUES (?,?,?,?,?,?)', (
                    file_id, import_job_uuid, path, file.content_type or os.path.splitext(file.filename)[1].lstrip('.').lower(), len(contents), now_iso()
                ))
            except Exception:
                pass
        conn.commit()
        return {'import_job_id': import_job_uuid, 'legacy_job_id': jid}
    finally:
        conn.close()



@router.post('/import/parse')
def parse_job_v3(payload: Dict[str, Any] = Body(...), sheet: Optional[str] = None, max_preview: int = 200, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    import_job_id = payload.get('import_job_id')
    if not import_job_id:
        raise HTTPException(status_code=400, detail='missing import_job_id')
    conn = db.connect()
    try:
        cur = conn.cursor()
        # some older local DBs may not have the `size_bytes` column on import_file
        # so only select the stored_path when present.
        # Attempt to locate the stored file path. Different schema versions
        # store provenance under different columns; try several fallbacks.
        f = None
        try:
            cur.execute('SELECT stored_path FROM import_file WHERE import_job_id=? ORDER BY uploaded_at DESC LIMIT 1', (import_job_id,))
            f = cur.fetchone()
        except Exception:
            f = None

        # If not found, try matching by filename/sha from import_job_v3 or fall back
        if not f:
            try:
                cur.execute('SELECT filename, file_sha256 FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
                j = cur.fetchone()
            except Exception:
                j = None
            fname = None
            fsha = None
            if j:
                try:
                    fname = j['filename']
                except Exception:
                    try:
                        fname = j[0]
                    except Exception:
                        fname = None
                try:
                    fsha = j['file_sha256']
                except Exception:
                    try:
                        fsha = j[1]
                    except Exception:
                        fsha = None
            if fname:
                try:
                    cur.execute('SELECT stored_path FROM import_file WHERE original_filename=? ORDER BY uploaded_at DESC LIMIT 1', (fname,))
                    f = cur.fetchone()
                except Exception:
                    f = None
            if not f and fsha:
                try:
                    cur.execute('SELECT stored_path FROM import_file WHERE sha256=? ORDER BY uploaded_at DESC LIMIT 1', (fsha,))
                    f = cur.fetchone()
                except Exception:
                    f = None
            # last-resort: take most recent stored_path
            if not f:
                try:
                    cur.execute('SELECT stored_path FROM import_file ORDER BY uploaded_at DESC LIMIT 1')
                    f = cur.fetchone()
                except Exception:
                    f = None

        if not f:
            # Try to locate a legacy numeric import_job folder where the file
            # was stored. The legacy `import_job` table uses numeric ids and
            # `job_{id}` folders — look up by file sha or filename if possible.
            try:
                # attempt to read v3 job filename/sha
                cur.execute('SELECT filename, file_sha256 FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
                j = cur.fetchone()
            except Exception:
                j = None
            legacy_found = None
            if j:
                try:
                    fname = j['filename'] if hasattr(j, 'keys') and 'filename' in j.keys() else (j[0] if len(j) > 0 else None)
                except Exception:
                    fname = None
                try:
                    fsha = j['file_sha256'] if hasattr(j, 'keys') and 'file_sha256' in j.keys() else (j[1] if len(j) > 1 else None)
                except Exception:
                    fsha = None
                try:
                    if fname:
                        cur.execute('SELECT id FROM import_job WHERE filename_original=? OR filename=? LIMIT 1', (fname, fname))
                        row = cur.fetchone()
                        if row:
                            legacy_found = row['id'] if hasattr(row, 'keys') and 'id' in row.keys() else row[0]
                    if not legacy_found and fsha:
                        cur.execute('SELECT id FROM import_job WHERE sha256_hash=? OR file_hash=? LIMIT 1', (fsha, fsha))
                        row = cur.fetchone()
                        if row:
                            legacy_found = row['id'] if hasattr(row, 'keys') and 'id' in row.keys() else row[0]
                except Exception:
                    legacy_found = None
            if legacy_found:
                uploads_base = os.getenv('TAAIP_UPLOAD_DIR', 'services/api/.data/imports')
                job_dir = os.path.join(uploads_base, f'job_{legacy_found}')
                if os.path.isdir(job_dir):
                    candidates = [os.path.join(job_dir, f) for f in os.listdir(job_dir) if f.startswith('original')]
                    if candidates:
                        path = candidates[0]
                    else:
                        path = None
                else:
                    path = None
            else:
                raise HTTPException(status_code=404, detail='uploaded file not found')
        else:
            # normalize fetched row to obtain stored_path
            try:
                path = f['stored_path'] if hasattr(f, 'keys') and 'stored_path' in f.keys() else f[0]
            except Exception:
                # final fallback: string value
                path = str(f)
        if not os.path.isfile(path):
            raise HTTPException(status_code=500, detail='uploaded file missing on disk')
        ext = os.path.splitext(path)[1].lower()
        # content sniff: allow CSVs without a .csv extension to be parsed
        if ext not in ('.csv', '.txt', '.json', '.xlsx', '.xls', '.sql'):
            try:
                with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                    sample = fh.read(2048)
                if '\n' in sample and ',' in sample:
                    ext = '.csv'
                elif sample.strip().startswith('{') or sample.strip().startswith('['):
                    ext = '.json'
            except Exception:
                pass
        preview = []
        columns = []
        # reuse parsing logic from existing function where possible
        if ext in ('.csv', '.txt'):
            with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                reader = csv.DictReader(fh)
                for i, r in enumerate(reader):
                    if i >= max_preview: break
                    preview.append({k: (v if v is not None else '') for k, v in r.items()})
            columns = list(preview[0].keys()) if preview else []
        elif ext == '.json':
            with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                data = json.load(fh)
                if isinstance(data, list):
                    preview = data[:max_preview]
                    columns = sorted(list(set().union(*(d.keys() for d in preview))) ) if preview else []
                elif isinstance(data, dict):
                    def flatten(d, parent=''):
                        items = {}
                        for k, v in d.items():
                            key = f"{parent}.{k}" if parent else k
                            if isinstance(v, dict):
                                items.update(flatten(v, key))
                            else:
                                items[key] = v
                        return items
                    flat = flatten(data)
                    preview = [flat]
                    columns = list(flat.keys())
        elif ext in ('.xlsx', '.xls'):
            try:
                import openpyxl as _openpyxl
            except Exception:
                raise HTTPException(status_code=202, detail='xlsx_parser_missing')
            wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_name = sheet or wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                header = []
            for i, r in enumerate(rows):
                if i >= max_preview: break
                rowobj = {str(header[j]) if header and j < len(header) and header[j] is not None else f'col_{j}': (r[j] if j < len(r) else None) for j in range(len(r))}
                preview.append(rowobj)
            columns = list(preview[0].keys()) if preview else []
        else:
            if ext == '.sql':
                # safe SQL parse: do NOT execute. Provide preview of first N non-empty lines and detect INSERT targets
                with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                    lines = [l.strip() for l in fh.readlines() if l.strip()]
                preview_lines = lines[:max_preview]
                preview = [{'sql': l} for l in preview_lines]
                columns = ['sql']
            else:
                raise HTTPException(status_code=400, detail='unsupported_file_type')

        # store preview in import_job_v3 notes (simple) and update row_count
        # also persist dataset_key if provided by the caller so downstream
        # commit logic knows which dataset to target without requiring legacy fallbacks
        try:
            ds = payload.get('dataset_key') if payload and isinstance(payload, dict) else None
        except Exception:
            ds = None
        if ds:
            cur.execute('UPDATE import_job_v3 SET dataset_key=?, row_count=? , status=?, notes=? WHERE id=?', (ds, len(preview), 'parsed', json.dumps({'columns': columns, 'preview_sample_count': len(preview)}), import_job_id))
        else:
            cur.execute('UPDATE import_job_v3 SET row_count=? , status=?, notes=? WHERE id=?', (len(preview), 'parsed', json.dumps({'columns': columns, 'preview_sample_count': len(preview)}), import_job_id))
        # also store preview rows into imported_rows (fallback provenance)
        for i, r in enumerate(preview):
            cur.execute('INSERT INTO imported_rows(import_job_id, target_domain, row_json, created_at) VALUES (?,?,?,?)', (import_job_id, payload.get('dataset_key','generic'), json.dumps(r), now_iso()))
        conn.commit()
        return {'status':'ok', 'import_job_id': import_job_id, 'columns': columns, 'preview_rows': preview[:50], 'row_count': len(preview)}
    finally:
        conn.close()


@router.post('/import/preview')
def preview_v3(payload: Dict[str, Any] = Body(...)):
    """Compatibility alias: POST /api/import/preview { import_job_id: '<uuid>' }
    Delegates to v3 parse logic (`/api/import/parse`)."""
    return parse_job_v3(payload)


@router.post('/import/{import_job_id}/parse')
@router.get('/import/{import_job_id}/parse')
@router.post('/import/{import_job_id}/parse/')
@router.get('/import/{import_job_id}/parse/')
def parse_job_compat(import_job_id: str, sheet: Optional[str] = None, max_preview: int = 200, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    """Compatibility alias: POST/GET /api/import/{id}/parse (with or without trailing slash)
    Delegates to /api/import/parse v3 logic. Accepts both POST (body) and GET (query param) callers.
    """
    return parse_job_v3({ 'import_job_id': import_job_id })


@router.post('/import/map')
def map_v3(payload: Dict[str, Any] = Body(...), mapping = None, dataset_key: Optional[str] = None, source_system: Optional[str] = None, scope_org: Optional[str] = None, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    import_job_id = payload.get('import_job_id') if payload else None
    mapping = payload.get('mapping') if payload else mapping
    dataset_key = payload.get('dataset_key') if payload else dataset_key
    source_system = payload.get('source_system') if payload else source_system
    scope_org = payload.get('scope_org_unit_id') if payload else scope_org
    if not import_job_id:
        raise HTTPException(status_code=400, detail='missing import_job_id')

    # Allow empty mapping {} as a signal to ask the server to auto-map fields
    # (client-side Auto-map sends {} when no client-side matches exist).
    if mapping is None:
        raise HTTPException(status_code=400, detail='missing fields')

    conn = db.connect()
    try:
        cur = conn.cursor()
        # If client requested server-side auto-mapping (empty mapping provided),
        # attempt to compute a canonical mapping from parsed preview columns.
        try:
            if isinstance(mapping, dict) and len(mapping) == 0:
                # Read parsed columns from import_job_v3.notes where parse stored preview info
                cur.execute('SELECT notes, dataset_key FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
                j = cur.fetchone()
                cols = []
                dskey = dataset_key
                if j:
                    try:
                        # normalize row
                        if hasattr(j, 'keys'):
                            notes_val = j['notes'] if 'notes' in j.keys() else None
                            dskey = dskey or (j['dataset_key'] if 'dataset_key' in j.keys() else dskey)
                        else:
                            notes_val = j[0]
                    except Exception:
                        notes_val = None
                    try:
                        if notes_val:
                            nj = json.loads(notes_val)
                            cols = nj.get('columns', []) if isinstance(nj, dict) else []
                    except Exception:
                        cols = []

                # Candidate target fields across datasets (kept conservative)
                candidate_targets = ['org_unit_id','date_key','metric_key','metric_value','event_name','start_date','end_date','city','state']
                auto_map = {}
                lowers = {c.lower(): c for c in cols}
                for t in candidate_targets:
                    if t in cols:
                        auto_map[t] = t
                    else:
                        # allow matching on lowercase equivalence
                        if t.lower() in lowers:
                            auto_map[t] = lowers[t.lower()]

                # If we found a non-empty auto_map, persist it as the mapping
                if auto_map:
                    mapping = auto_map
                    # Guess dataset_key when client used 'auto' so commit uses correct loader
                    if not dataset_key or dataset_key in ('auto', 'generic', None):
                        if any(k in mapping for k in ('metric_key','metric_value','org_unit_id')):
                            dataset_key = 'production'
                        elif any(k in mapping for k in ('event_name','start_date','end_date')):
                            dataset_key = 'events'
                        else:
                            dataset_key = 'generic'
                # else leave mapping as {} and let later validation respond with helpful message
        except Exception:
            # on any error while auto-mapping, proceed with original mapping value
            pass
        try:
            # reject completely missing mapping (None) above; here mapping may be {} or populated
            if isinstance(mapping, dict) and len(mapping) == 0:
                # explicit server-side rejection with clearer diagnostic
                raise HTTPException(status_code=400, detail='missing fields: mapping empty after auto-map')

            map_id = uuid.uuid4().hex
            cur.execute('INSERT INTO import_column_map(id, import_job_id, mapping_json, created_at) VALUES (?,?,?,?)', (map_id, import_job_id, json.dumps(mapping), now_iso()))
            cur.execute('UPDATE import_job_v3 SET dataset_key=?, source_system=?, scope_org_unit_id=?, status=?, updated_at=? WHERE id=?', (dataset_key, source_system, scope_org, 'mapped', now_iso(), import_job_id))
            conn.commit()
            return {'status':'ok', 'import_job_id': import_job_id}
        except Exception as e:
            # Fallback: attempt to find a legacy numeric import_job that corresponds
            # to this v3 job (matching filename or sha) and apply mapping there.
            try:
                cur.execute('SELECT filename, file_sha256 FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
                j = cur.fetchone()
                legacy_id = None
                if j:
                    # convert sqlite3.Row/dict-like to tuple/dict safely
                    try:
                        filename = j['filename'] if 'filename' in j.keys() else (j[0] if len(j) > 0 else None)
                    except Exception:
                        filename = j[0] if j and len(j) > 0 else None
                    try:
                        sha = j['file_sha256'] if 'file_sha256' in j.keys() else (j[1] if len(j) > 1 else None)
                    except Exception:
                        sha = j[1] if j and len(j) > 1 else None

                    if filename:
                        cur.execute('SELECT id FROM import_job WHERE filename_original=? OR filename=? LIMIT 1', (filename, filename))
                        found = cur.fetchone()
                        if found:
                            legacy_id = found['id'] if 'id' in found.keys() else found[0]
                    if not legacy_id and sha:
                        cur.execute('SELECT id FROM import_job WHERE sha256_hash=? OR file_hash=? LIMIT 1', (sha, sha))
                        found = cur.fetchone()
                        if found:
                            legacy_id = found['id'] if 'id' in found.keys() else found[0]

                if legacy_id:
                    # store mapping into legacy import_job mapping_json and mapping template
                    try:
                        cur.execute('INSERT INTO import_mapping_template(name, target_domain, mapping_json, created_by, created_at) VALUES (?,?,?,?,?)', (
                            mapping.get('name', f'map_{legacy_id}'), mapping.get('target_domain','generic'), json.dumps(mapping), mapping.get('created_by'), now_iso()
                        ))
                    except Exception:
                        pass
                    try:
                        cur.execute('UPDATE import_job SET status=?, target_domain=?, mapping_json=?, updated_at=? WHERE id=?', ('validating', mapping.get('target_domain','generic'), json.dumps(mapping), now_iso(), legacy_id))
                        conn.commit()
                    except Exception:
                        pass
                    return {'status':'ok', 'import_job_id': import_job_id, 'legacy_job_id': legacy_id, 'mapped_via_legacy': True}
            except Exception:
                pass
            # Re-raise original for visibility if fallback didn't succeed
            raise
    finally:
        conn.close()


@router.post('/import/validate')
def validate_v3(payload: Dict[str, Any] = Body(...), allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    import_job_id = payload.get('import_job_id')
    if not import_job_id:
        raise HTTPException(status_code=400, detail='missing import_job_id')
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT mapping_json FROM import_column_map WHERE import_job_id=? ORDER BY created_at DESC LIMIT 1', (import_job_id,))
        mrow = cur.fetchone()
        if mrow and not hasattr(mrow, 'keys'):
            try:
                mrow = {d[0]: mrow[i] for i, d in enumerate(cur.description)}
            except Exception:
                pass
        mapping = json.loads(mrow.get('mapping_json')) if mrow and isinstance(mrow, dict) and mrow.get('mapping_json') else {}
        # dataset_key is stored on the import_job_v3 record; fall back to payload if absent
        cur.execute('SELECT dataset_key FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
        jrow = cur.fetchone()
        if jrow and not hasattr(jrow, 'keys'):
            try:
                jrow = {d[0]: jrow[i] for i, d in enumerate(cur.description)}
            except Exception:
                pass
        dataset = jrow.get('dataset_key') if jrow and isinstance(jrow, dict) and jrow.get('dataset_key') else (payload.get('dataset_key') if payload else None)
        # fetch imported_rows for this job
        cur.execute('SELECT id, row_json FROM imported_rows WHERE import_job_id=? LIMIT 1000', (import_job_id,))
        rows = cur.fetchall()
        if rows and not hasattr(rows[0], 'keys'):
            try:
                desc = cur.description
                rows = [{desc[i][0]: r[i] for i in range(len(desc))} for r in rows]
            except Exception:
                pass
        errors = 0
        sample_errors = []
        for i, r in enumerate(rows):
            row = json.loads(r['row_json']) if isinstance(r['row_json'], str) else r['row_json']
            # basic validations depending on dataset
            if dataset == 'production':
                # require org_unit_id, date (YYYY-MM-DD), metric_key, metric_value
                org = row.get(mapping.get('org_unit_id')) if mapping.get('org_unit_id') else row.get('org_unit_id') or row.get('org_unit')
                date = row.get(mapping.get('date_key')) if mapping.get('date_key') else row.get('date') or row.get('date_key')
                metric = row.get(mapping.get('metric_key')) if mapping.get('metric_key') else row.get('metric_key')
                value = row.get(mapping.get('metric_value')) if mapping.get('metric_value') else row.get('metric_value')
                # org check
                if not org:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'org_unit', 'missing org identifier', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'org_unit', 'message':'missing org identifier'})
                # date check
                try:
                    if date:
                        _ = datetime.datetime.strptime(str(date)[:10], '%Y-%m-%d')
                    else:
                        raise Exception()
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'date', 'invalid date', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'date', 'message':'invalid date'})
                # numeric value
                try:
                    float(value)
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'metric_value', 'not numeric', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'metric_value', 'message':'not numeric'})
            elif dataset == 'marketing':
                # check date and org
                date = row.get(mapping.get('date_key')) if mapping.get('date_key') else row.get('date')
                if not date:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'date', 'missing date', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'date', 'message':'missing date'})
            elif dataset == 'funnel':
                org = row.get(mapping.get('org_unit_id')) if mapping.get('org_unit_id') else row.get('org_unit_id') or row.get('org_unit')
                date = row.get(mapping.get('date_key')) if mapping.get('date_key') else row.get('date') or row.get('date_key')
                stage = row.get(mapping.get('stage')) if mapping.get('stage') else row.get('stage')
                count = row.get(mapping.get('count_value')) if mapping.get('count_value') else row.get('count_value')
                if not org:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'org_unit', 'missing org identifier', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'org_unit', 'message':'missing org identifier'})
                try:
                    if date:
                        _ = datetime.datetime.strptime(str(date)[:10], '%Y-%m-%d')
                    else:
                        raise Exception()
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'date', 'invalid date', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'date', 'message':'invalid date'})
                try:
                    float(count)
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'count_value', 'not numeric', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'count_value', 'message':'not numeric'})
            elif dataset == 'org_units':
                # require id and type
                uid = row.get(mapping.get('id')) if mapping.get('id') else row.get('id')
                utype = row.get(mapping.get('type')) if mapping.get('type') else row.get('type')
                if not uid or not utype:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'missing id or type', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'missing id or type'})
            elif dataset in ('expenses', 'expense'):
                # require amount numeric and at least one of project_id/event_id/org_unit_id
                amt = row.get(mapping.get('amount')) if mapping.get('amount') else row.get('amount')
                pid = row.get(mapping.get('project_id')) if mapping.get('project_id') else row.get('project_id')
                eid = row.get(mapping.get('event_id')) if mapping.get('event_id') else row.get('event_id')
                org = row.get(mapping.get('org_unit_id')) if mapping.get('org_unit_id') else row.get('org_unit_id')
                try:
                    float(amt)
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'amount', 'invalid amount', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'amount', 'message':'invalid amount'})
                if not (pid or eid or org):
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'missing project_id/event_id/org_unit', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'missing project_id/event_id/org_unit'})
            elif dataset in ('budget_line_item', 'budget'):
                fyv = row.get(mapping.get('fy')) if mapping.get('fy') else row.get('fy')
                org = row.get(mapping.get('org_unit_id')) if mapping.get('org_unit_id') else row.get('org_unit_id')
                amt = row.get(mapping.get('amount')) if mapping.get('amount') else row.get('amount')
                if fyv is None or org is None:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'missing fy or org_unit_id', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'missing fy or org_unit_id'})
                try:
                    float(amt)
                except Exception:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, 'amount', 'invalid amount', now_iso()))
                    sample_errors.append({'row': i+1, 'field':'amount', 'message':'invalid amount'})
            elif dataset == 'projects':
                pid = row.get(mapping.get('project_id')) if mapping.get('project_id') else row.get('project_id')
                title = row.get(mapping.get('title')) if mapping.get('title') else row.get('title')
                if not pid or not title:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'missing project_id or title', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'missing project_id or title'})
            elif dataset == 'events':
                eid = row.get(mapping.get('event_id')) if mapping.get('event_id') else row.get('event_id')
                name = row.get(mapping.get('name')) if mapping.get('name') else row.get('name')
                if not eid or not name:
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'missing event_id or name', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'missing event_id or name'})
            else:
                # generic checks: non-empty row
                if not any(row.values()):
                    errors += 1
                    cur.execute('INSERT INTO import_error(id, import_job_id, row_index, field, message, created_at) VALUES (?,?,?,?,?,?)', (uuid.uuid4().hex, import_job_id, i+1, None, 'empty row', now_iso()))
                    sample_errors.append({'row': i+1, 'message':'empty row'})

        cur.execute('UPDATE import_job_v3 SET error_count=?, status=?, updated_at=? WHERE id=?', (errors, ('validated' if errors==0 else 'validated_with_errors'), now_iso(), import_job_id))
        conn.commit()
        return {'status':'ok', 'import_job_id': import_job_id, 'error_count': errors, 'sample_errors': sample_errors[:50]}
    finally:
        conn.close()


@router.post('/import/commit')
def commit_v3(payload: Dict[str, Any] = Body(...), allowed_orgs: Optional[list] = Depends(require_scope('STATION')), current_user: Dict = Depends(get_current_user)):
    import_job_id = payload.get('import_job_id')
    mode = payload.get('mode', 'append')
    debug = bool(payload.get('debug'))
    if not import_job_id:
        raise HTTPException(status_code=400, detail='missing import_job_id')
    conn = db.connect()
    try:
        cur = conn.cursor()
        # Try v3 import_job first
        cur.execute('SELECT dataset_key, source_system, scope_org_unit_id FROM import_job_v3 WHERE id=?', (import_job_id,))
        job = cur.fetchone()
        rows = []
        legacy_mode = False
        if job:
            # normalize job whether it's a sqlite3.Row or a plain tuple
            if hasattr(job, 'keys'):
                try:
                    dataset = job['dataset_key']
                except Exception:
                    dataset = None
                try:
                    source_system = job['source_system']
                except Exception:
                    source_system = None
                try:
                    scope_org = job['scope_org_unit_id']
                except Exception:
                    scope_org = None
            else:
                try:
                    dataset = job[0]
                except Exception:
                    dataset = None
                try:
                    source_system = job[1]
                except Exception:
                    source_system = None
                try:
                    scope_org = job[2]
                except Exception:
                    scope_org = None
            # fetch v3 stored preview rows
            cur.execute('SELECT row_json FROM imported_rows WHERE import_job_id=?', (import_job_id,))
            rows = cur.fetchall()
            try:
                print(f"commit_v3: fetched {len(rows)} imported_rows for import_job_id={import_job_id}")
            except Exception:
                pass
            # normalize rows to mapping-like dicts if DB driver returned tuples
            if rows and not hasattr(rows[0], 'keys'):
                try:
                    desc = cur.description
                    rows = [{desc[i][0]: r[i] for i in range(len(desc))} for r in rows]
                except Exception:
                    pass
            # load mapping if present
            cur.execute('SELECT mapping_json FROM import_column_map WHERE import_job_id=? ORDER BY created_at DESC LIMIT 1', (import_job_id,))
            mrow = cur.fetchone()
            if mrow and not hasattr(mrow, 'keys'):
                try:
                    mrow = {d[0]: mrow[i] for i, d in enumerate(cur.description)}
                except Exception:
                    pass
            # mrow may be a sqlite3.Row (has keys() but no get()), a dict, or None
            mapping_json_val = None
            if mrow:
                if hasattr(mrow, 'keys'):
                    try:
                        mapping_json_val = mrow['mapping_json'] if 'mapping_json' in mrow.keys() else None
                    except Exception:
                        mapping_json_val = None
                elif isinstance(mrow, dict):
                    mapping_json_val = mrow.get('mapping_json')
            mapping = json.loads(mapping_json_val) if mapping_json_val else {}
            # normalize mapping to a simple field_map dict when mapping contains a 'field_map' wrapper
            field_map = mapping.get('field_map') if isinstance(mapping, dict) and mapping.get('field_map') else mapping
            # If this is a pure v3 flow (not legacy) we require a mapping or a concrete dataset_key
            if not legacy_mode:
                if (not mapping or mapping == {}) and (not dataset or dataset in ('generic', None)):
                    raise HTTPException(status_code=400, detail='mapping_required: provide mapping or dataset_key before commit')
        else:
            # fallback: support legacy numeric import_job id
            try:
                legacy_id = int(import_job_id)
            except Exception:
                legacy_id = None
            if legacy_id is not None:
                cur.execute('SELECT id, target_domain, source_system, filename FROM import_job WHERE id=?', (legacy_id,))
                ljob = cur.fetchone()
                if ljob:
                    legacy_mode = True
                    if hasattr(ljob, 'keys'):
                        dataset = ljob.get('target_domain')
                        source_system = ljob.get('source_system')
                    else:
                        try:
                            dataset = ljob[1] if len(ljob) > 1 else 'generic'
                        except Exception:
                            dataset = 'generic'
                        try:
                            source_system = ljob[2] if len(ljob) > 2 else None
                        except Exception:
                            source_system = None
                    scope_org = None
                    # fetch legacy imported_rows where import_job_id is numeric
                    cur.execute('SELECT row_json FROM imported_rows WHERE import_job_id=?', (legacy_id,))
                    rows = cur.fetchall()
                    if rows and not hasattr(rows[0], 'keys'):
                        try:
                            desc = cur.description
                            rows = [{desc[i][0]: r[i] for i in range(len(desc))} for r in rows]
                        except Exception:
                            pass
                else:
                    raise HTTPException(status_code=404, detail='job not found')
            else:
                raise HTTPException(status_code=404, detail='job not found')
        committed = 0
        diagnostics = []
        for r in rows:
            row = json.loads(r['row_json']) if isinstance(r.get('row_json'), str) else r.get('row_json')
            # Defensive normalization: some legacy/driver combinations may
            # return a sequence or a dict with a single 'value' key that
            # contains the original tuple (row_json, target_domain, created_at).
            # Attempt to recover a JSON string if present as the first element.
            try:
                if not isinstance(row, dict):
                    if isinstance(row, (list, tuple)) and len(row) > 0 and isinstance(row[0], str):
                        try:
                            row = json.loads(row[0])
                        except Exception:
                            row = {'value': list(row)}
                else:
                    # dict with single 'value' key wrapping the tuple
                    if set(row.keys()) == {'value'} and isinstance(row['value'], (list, tuple)):
                        v = row['value']
                        if len(v) > 0 and isinstance(v[0], str):
                            try:
                                row = json.loads(v[0])
                            except Exception:
                                row = {'value': list(v)}
            except Exception:
                pass
            try:
                print(f"commit_v3: processing row: {row}")
            except Exception:
                pass
            if dataset == 'production':
                # use mapping when provided
                org = row.get(field_map.get('org_unit_id')) if field_map and field_map.get('org_unit_id') else (row.get('org_unit_id') or row.get('org_unit'))
                date = row.get(field_map.get('date_key')) if field_map and field_map.get('date_key') else (row.get('date') or row.get('date_key'))
                metric = row.get(field_map.get('metric_key')) if field_map and field_map.get('metric_key') else row.get('metric_key')
                val = row.get(field_map.get('metric_value')) if field_map and field_map.get('metric_value') else row.get('metric_value')
                if not org or not date or metric is None or val is None:
                    continue
                fid = uuid.uuid4().hex
                if mode and str(mode).startswith('replace'):
                    # archive any existing active rows matching the business key
                    try:
                        cur.execute('UPDATE fact_production SET record_status=?, archived_at=? WHERE org_unit_id=? AND date_key=? AND metric_key=? AND (record_status IS NULL OR record_status="active")', ('archived', now_iso(), str(org), str(date)[:10], str(metric)))
                    except Exception:
                        pass
                cur.execute('INSERT OR REPLACE INTO fact_production(id, org_unit_id, date_key, metric_key, metric_value, source_system, import_job_id, created_at) VALUES (?,?,?,?,?,?,?,?)', (fid, str(org), str(date)[:10], str(metric), float(val), source_system, import_job_id, now_iso()))
                committed += 1
            elif dataset == 'marketing':
                fid = uuid.uuid4().hex
                org = row.get(field_map.get('org_unit_id')) if field_map and field_map.get('org_unit_id') else (row.get('org_unit_id') or row.get('org_unit'))
                date = row.get(field_map.get('date_key')) if field_map and field_map.get('date_key') else (row.get('date') or row.get('date_key'))
                if mode and str(mode).startswith('replace'):
                    try:
                        cur.execute('UPDATE fact_marketing SET record_status=?, archived_at=? WHERE org_unit_id=? AND date_key=? AND (record_status IS NULL OR record_status="active")', ('archived', now_iso(), str(org), str(date)[:10] if date else None))
                    except Exception:
                        pass
                cur.execute('INSERT OR REPLACE INTO fact_marketing(id, org_unit_id, date_key, campaign, channel, impressions, engagements, clicks, conversions, cost, source_system, import_job_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)', (
                    fid, str(org), str(date)[:10] if date else None, row.get(field_map.get('campaign') if field_map and field_map.get('campaign') else 'campaign') if field_map else row.get('campaign'), row.get(field_map.get('channel') if field_map and field_map.get('channel') else 'channel') if field_map else row.get('channel'), float(row.get(field_map.get('impressions') if field_map and field_map.get('impressions') else 'impressions') or 0), float(row.get(field_map.get('engagements') if field_map and field_map.get('engagements') else 'engagements') or 0), float(row.get(field_map.get('clicks') if field_map and field_map.get('clicks') else 'clicks') or 0), float(row.get(field_map.get('conversions') if field_map and field_map.get('conversions') else 'conversions') or 0), float(row.get(field_map.get('cost') if field_map and field_map.get('cost') else 'cost') or 0), source_system, import_job_id, now_iso()
                ))
                committed += 1
            elif dataset == 'event_performance' or dataset == 'event_metrics':
                # map into event_metrics table
                fid = uuid.uuid4().hex
                event_id = row.get(field_map.get('event_id')) if field_map and field_map.get('event_id') else row.get('event_id') or row.get('event')
                try:
                    impressions = int(row.get(field_map.get('impressions')) or row.get('impressions') or 0)
                except Exception:
                    impressions = 0
                try:
                    engagements = int(row.get(field_map.get('engagements')) or row.get('engagements') or 0)
                except Exception:
                    engagements = 0
                try:
                    captured_at = row.get(field_map.get('captured_at')) if field_map and field_map.get('captured_at') else row.get('captured_at') or row.get('date')
                except Exception:
                    captured_at = None
                cur.execute('INSERT OR REPLACE INTO event_metrics(id, event_id, impressions, engagements, leads, appts_made, appts_conducted, contracts, accessions, other_json, captured_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)', (
                    fid, event_id, impressions, engagements, int(row.get('leads') or 0), int(row.get('appts_made') or 0), int(row.get('appts_conducted') or 0), int(row.get('contracts') or 0), int(row.get('accessions') or 0), json.dumps(row), captured_at
                ))
                committed += 1
            elif dataset == 'funnel' or dataset == 'leads':
                # map into fact_funnel table
                fid = uuid.uuid4().hex
                org = row.get(field_map.get('org_unit_id')) if field_map and field_map.get('org_unit_id') else (row.get('org_unit_id') or row.get('org_unit'))
                date = row.get(field_map.get('date_key')) if field_map and field_map.get('date_key') else (row.get('date') or row.get('date_key'))
                stage = row.get(field_map.get('stage')) if field_map and field_map.get('stage') else row.get('stage')
                count = row.get(field_map.get('count_value')) if field_map and field_map.get('count_value') else row.get('count_value')
                if debug:
                    diagnostics.append({'target_table': 'fact_funnel', 'cols': ['org_unit_id','date_key','stage','count_value'], 'vals': [org, date, stage, count]})
                try:
                    if mode and str(mode).startswith('replace'):
                        try:
                            cur.execute('UPDATE fact_funnel SET record_status=?, archived_at=? WHERE org_unit_id=? AND date_key=? AND (record_status IS NULL OR record_status="active")', ('archived', now_iso(), str(org), str(date)[:10] if date else None))
                        except Exception as e:
                            if debug:
                                diagnostics.append({'target_table': 'fact_funnel', 'action': 'archive', 'error': str(e)})
                    cur.execute('INSERT OR REPLACE INTO fact_funnel(id, org_unit_id, date_key, stage, count_value, source_system, import_job_id, created_at) VALUES (?,?,?,?,?,?,?,?)', (
                        fid, str(org) if org is not None else None, str(date)[:10] if date else None, stage, float(count) if count not in (None, '') else 0, source_system, import_job_id, now_iso()
                    ))
                    committed += 1
                except Exception as e:
                    # record diagnostic and continue; do not silently swallow
                    if debug:
                        diagnostics.append({'target_table': 'fact_funnel', 'cols': ['org_unit_id','date_key','stage','count_value'], 'vals':[org,date,stage,count], 'error': str(e)})
                    else:
                        diagnostics.append({'target_table': 'fact_funnel', 'error': 'insert_failed'})
            else:
                # fallback: keep in imported_rows only
                continue
        # mark provenance rows processed (best-effort)
        try:
            uname = current_user.get('username') if isinstance(current_user, dict) else None
            cur.execute('UPDATE imported_rows SET processed=1, processed_at=?, processed_by=? WHERE import_job_id=?', (now_iso(), uname or 'system', import_job_id))
        except Exception:
            pass

        # Update provenance status on the correct table
        if legacy_mode:
            try:
                cur.execute('UPDATE import_job SET status=?, row_count_detected=? WHERE id=?', ('committed', committed, legacy_id))
            except Exception:
                pass
        else:
            cur.execute('UPDATE import_job_v3 SET status=?, row_count=?, updated_at=? WHERE id=?', ('committed', committed, now_iso(), import_job_id))
        conn.commit()
        resp = {'status':'ok', 'import_job_id': import_job_id, 'committed_rows': committed}
        if debug:
            resp['diagnostics'] = diagnostics
        return resp
    finally:
        conn.close()


@router.get('/import/jobs')
def list_import_jobs(limit: int = 100, current_user: Optional[dict] = Depends(get_current_user)):
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT * FROM import_job_v3 ORDER BY created_at DESC LIMIT ?', (limit,))
        return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


@router.get('/import/jobs/{import_job_id}')
def get_import_job(import_job_id: str, current_user: Optional[dict] = Depends(get_current_user)):
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT * FROM import_job_v3 WHERE id=?', (import_job_id,))
        job = cur.fetchone()
        if not job:
            raise HTTPException(status_code=404, detail='not found')
        cur.execute('SELECT * FROM import_file WHERE import_job_id=? ORDER BY uploaded_at DESC', (import_job_id,))
        files = [dict(r) for r in cur.fetchall()]
        cur.execute('SELECT * FROM import_column_map WHERE import_job_id=? ORDER BY created_at DESC', (import_job_id,))
        maps = [dict(r) for r in cur.fetchall()]
        cur.execute('SELECT * FROM import_error WHERE import_job_id=? ORDER BY created_at LIMIT 100', (import_job_id,))
        errs = [dict(r) for r in cur.fetchall()]
        return {'job': dict(job), 'files': files, 'mappings': maps, 'errors': errs}
    finally:
        conn.close()


@router.get('/import/templates/{dataset_key}')
def get_template(dataset_key: str):
    # very small in-code templates
    templates = {
        'production': {
            'csv_header': 'org_unit_id,date,metric_key,metric_value',
            'fields': ['org_unit_id','date','metric_key','metric_value']
        },
        'marketing': {
            'csv_header': 'org_unit_id,date,campaign,channel,impressions,engagements,clicks,conversions,cost',
            'fields': ['org_unit_id','date','campaign','channel','impressions','engagements','clicks','conversions','cost']
        },
        'org_units': {
            'csv_header': 'id,name,type,parent_id,rsid,uic,city,state,zip',
            'fields': ['id','name','type','parent_id','rsid','uic','city','state','zip']
        }
        ,
        'funnel': {
            'csv_header': 'org_unit_id,date,stage,count_value',
            'fields': ['org_unit_id','date','stage','count_value']
        }
        ,
        'expenses': {
            'csv_header': 'project_id,event_id,fy,qtr,org_unit_id,station_id,funding_line,category,amount,spent_at,vendor,notes',
            'fields': ['project_id','event_id','fy','qtr','org_unit_id','station_id','funding_line','category','amount','spent_at','vendor','notes']
        },
        'budget_line_item': {
            'csv_header': 'org_unit_id,fy,qtr,event_id,category,description,amount,obligation_date,notes',
            'fields': ['org_unit_id','fy','qtr','event_id','category','description','amount','obligation_date','notes']
        },
        'projects': {
            'csv_header': 'project_id,title,org_unit_id,fy,qtr,funding_line,category,planned_cost,percent_complete,updated_at',
            'fields': ['project_id','title','org_unit_id','fy','qtr','funding_line','category','planned_cost','percent_complete','updated_at']
        },
        'events': {
            'csv_header': 'event_id,name,org_unit_id,fy,qtr,project_id,planned_cost,loe,start_dt,end_dt',
            'fields': ['event_id','name','org_unit_id','fy','qtr','project_id','planned_cost','loe','start_dt','end_dt']
        }
    }
    return templates.get(dataset_key, {'csv_header':'','fields':[]})


@router.post('/import/{job_id}/parse')
def parse_job(job_id: str, sheet: Optional[str] = None, max_preview: int = 200, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    """Legacy numeric job parser that also accepts v3 import_job_id (UUID string).
    If `job_id` looks numeric it follows the legacy parse flow; otherwise it
    delegates to the v3 `parse_job_v3` handler using the UUID import_job_id.
    """
    # detect legacy numeric id vs v3 uuid string
    try:
        legacy_id = int(job_id)
    except Exception:
        legacy_id = None
    if legacy_id is None:
        return parse_job_v3({ 'import_job_id': job_id })
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT id, filename_original, sha256_hash, file_type FROM import_job WHERE id=?', (job_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='import job not found')
        uploads_base = os.getenv('TAAIP_UPLOAD_DIR', 'services/api/.data/imports')
        job_dir = os.path.join(uploads_base, f'job_{job_id}')
        if not os.path.isdir(job_dir):
            raise HTTPException(status_code=500, detail='uploaded file missing')
        # expected stored as original.ext
        candidates = [os.path.join(job_dir, f) for f in os.listdir(job_dir) if f.startswith('original')]
        if not candidates:
            raise HTTPException(status_code=500, detail='uploaded file missing')
        path = candidates[0]
        ext = os.path.splitext(path)[1].lower()
        preview = []
        columns = []
        # CSV/TXT
        if ext in ('.csv', '.txt'):
            with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                reader = csv.DictReader(fh)
                for i, r in enumerate(reader):
                    if i >= max_preview: break
                    preview.append({k: (v if v is not None else '') for k, v in r.items()})
            columns = list(preview[0].keys()) if preview else []
        elif ext == '.json':
            with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                data = json.load(fh)
                if isinstance(data, list):
                    preview = data[:max_preview]
                    columns = sorted(list(set().union(*(d.keys() for d in preview))) ) if preview else []
                elif isinstance(data, dict):
                    # flatten nested dicts to dot notation
                    def flatten(d, parent=''):
                        items = {}
                        for k, v in d.items():
                            key = f"{parent}.{k}" if parent else k
                            if isinstance(v, dict):
                                items.update(flatten(v, key))
                            else:
                                items[key] = v
                        return items
                    flat = flatten(data)
                    preview = [flat]
                    columns = list(flat.keys())
                else:
                    preview = []
        elif ext in ('.xlsx', '.xls'):
            if openpyxl is None:
                try:
                    _update_import_job_status_fallback(cur, job_id, 'mapping_required', 'XLSX parsing requires openpyxl. Install it in the services/api venv.')
                except Exception:
                    try:
                        cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', ('mapping_required', now_iso(), job_id))
                    except Exception:
                        pass
                conn.commit()
                raise HTTPException(status_code=202, detail='mapping_required')
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_name = sheet or wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                header = []
            for i, r in enumerate(rows):
                if i >= max_preview: break
                rowobj = {str(header[j]) if header and j < len(header) and header[j] is not None else f'col_{j}': (r[j] if j < len(r) else None) for j in range(len(r))}
                preview.append(rowobj)
            columns = list(preview[0].keys()) if preview else []
        elif ext in ('.sql',):
            # attempt to parse simple INSERT INTO ... VALUES (...) statements
            with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                sql = fh.read()
            inserts = re.findall(r"INSERT\s+INTO\s+[^()]+\(([^)]+)\)\s+VALUES\s*\(([^;]+?)\)", sql, flags=re.IGNORECASE)
            if not inserts:
                # try parsing VALUES (...) without column list
                vals = re.findall(r"VALUES\s*\(([^;]+?)\)", sql, flags=re.IGNORECASE)
                if not vals:
                    try:
                        _update_import_job_status_fallback(cur, job_id, 'mapping_required', 'SQL file not in simple INSERT/VALUES form. Provide mapping.')
                    except Exception:
                        try:
                            cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', ('mapping_required', now_iso(), job_id))
                        except Exception:
                            pass
                    conn.commit()
                    raise HTTPException(status_code=202, detail='mapping_required')
                # produce preview rows with generic col_1..N
                for i, v in enumerate(vals[:max_preview]):
                    parts = [p.strip().strip("'\"") for p in v.split(',')]
                    preview.append({f'col_{j+1}': parts[j] if j < len(parts) else None for j in range(len(parts))})
                columns = list(preview[0].keys()) if preview else []
            else:
                # take first insert structure
                cols = [c.strip().strip('"\'') for c in inserts[0][0].split(',')]
                vals = inserts[0][1]
                parts = [p.strip().strip("'\"") for p in vals.split(',')]
                preview.append({cols[j] if j < len(cols) else f'col_{j+1}': parts[j] if j < len(parts) else None for j in range(len(parts))})
                columns = cols
        else:
            try:
                _update_import_job_status_fallback(cur, job_id, 'mapping_required', 'Parsing for this file type is not automated in local dev. Please provide mapping.')
            except Exception:
                try:
                    cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', ('mapping_required', now_iso(), job_id))
                except Exception:
                    pass
            conn.commit()
            raise HTTPException(status_code=202, detail='mapping_required')

        # store preview
        cur.execute('INSERT INTO import_job_preview(import_job_id, preview_json, columns_json, created_at) VALUES (?,?,?,?)', (job_id, json.dumps(preview), json.dumps(columns), now_iso()))
        cur.execute('UPDATE import_job SET status=?, row_count_detected=?, updated_at=? WHERE id=?', ('preview_ready', len(preview), now_iso(), job_id))
        conn.commit()
        return {'import_job_id': job_id, 'preview_rows': len(preview), 'columns': columns}
    finally:
        conn.close()


@router.get('/import/{job_id}')
def get_import(job_id: str, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    """Legacy numeric import getter that also accepts v3 import_job_id UUIDs.
    Numeric ids follow the legacy path; UUID strings are delegated to the
    v3 `get_import_job` helper which returns the v3 shaped job metadata.
    """
    try:
        legacy_id = int(job_id)
    except Exception:
        legacy_id = None
    if legacy_id is None:
        try:
            return get_import_job(job_id)
        except Exception:
            # Best-effort fallback when the local DB schema doesn't match
            # expectations (older schema). Return available v3 job row and
            # empty artifacts to allow the UI to continue rather than
            # crashing the whole request.
            conn = db.connect()
            try:
                cur = conn.cursor()
                cur.execute('SELECT * FROM import_job_v3 WHERE id=? LIMIT 1', (job_id,))
                j = cur.fetchone()
                if not j:
                    raise HTTPException(status_code=404, detail='not found')
                try:
                    job = dict(j)
                except Exception:
                    # sqlite3.Row fallback
                    try:
                        desc = cur.description
                        job = {desc[i][0]: j[i] for i in range(len(desc))}
                    except Exception:
                        job = {}
                return {'job': job, 'files': [], 'mappings': [], 'errors': []}
            finally:
                conn.close()
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT * FROM import_job WHERE id=?', (job_id,))
        job = cur.fetchone()
        if not job:
            raise HTTPException(status_code=404, detail='not found')
        cur.execute('SELECT preview_json, columns_json FROM import_job_preview WHERE import_job_id=? ORDER BY id DESC LIMIT 1', (job_id,))
        p = cur.fetchone()
        preview = json.loads(p['preview_json']) if p else []
        columns = json.loads(p['columns_json']) if p else []
        # fetch logs
        cur.execute('SELECT level, message, row_number, field_name, created_at FROM import_job_log WHERE import_job_id=? ORDER BY id', (job_id,))
        logs = [dict(r) for r in cur.fetchall()]
        return {'job': dict(job), 'preview': preview, 'columns': columns, 'logs': logs}
    finally:
        conn.close()


@router.post('/import/{job_id}/map')
def map_job(job_id: str, mapping: Dict, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    """Legacy numeric map endpoint that also accepts v3 import_job_id UUIDs.
    If given a UUID, delegate to the v3 `map_v3` handler.
    """
    try:
        legacy_id = int(job_id)
    except Exception:
        legacy_id = None
    if legacy_id is None:
        # If mapping not provided or empty, attempt a best-effort auto-map
        # by inspecting any parsed preview or job notes so the UI's quick
        # "Auto-map same-name" flow doesn't fail when the client raced
        # state updates.
        mg = mapping if mapping else None
        if not mg:
            try:
                # try v3 job notes for columns
                conn = db.connect()
                cur = conn.cursor()
                cur.execute('SELECT dataset_key, notes FROM import_job_v3 WHERE id=? LIMIT 1', (job_id,))
                j = cur.fetchone()
                cols = []
                ds = None
                if j:
                    try:
                        # sqlite3.Row may be dict-like
                        ds = j['dataset_key'] if 'dataset_key' in j.keys() else (j[0] if len(j)>0 else None)
                    except Exception:
                        try:
                            ds = j[0]
                        except Exception:
                            ds = None
                    try:
                        notes = j['notes'] if 'notes' in j.keys() else (j[1] if len(j)>1 else None)
                    except Exception:
                        notes = None
                    if notes:
                        try:
                            nj = json.loads(notes)
                            cols = nj.get('columns', []) if isinstance(nj, dict) else []
                        except Exception:
                            cols = []
                # fallback to import_job_preview table
                if not cols:
                    try:
                        cur.execute('SELECT columns_json FROM import_job_preview WHERE import_job_id=? ORDER BY id DESC LIMIT 1', (job_id,))
                        p = cur.fetchone()
                        if p:
                            try:
                                cols = json.loads(p['columns_json']) if 'columns_json' in p.keys() else json.loads(p[0])
                            except Exception:
                                cols = []
                    except Exception:
                        cols = []
            except Exception:
                cols = []
            finally:
                try: conn.close()
                except Exception: pass

            if cols:
                # build a field_map mapping target_field -> source_column
                # Prefer dataset-specific template fields, then common targets,
                # using case-insensitive, punctuation-stripped matching.
                try:
                    tmpl = get_template(ds) if ds else {}
                    tfields = tmpl.get('fields', []) if tmpl else []
                except Exception:
                    tfields = []

                def _norm(s):
                    try:
                        return re.sub(r'[^a-z0-9]', '', str(s).lower())
                    except Exception:
                        return str(s).lower()

                cols_norm = {c: _norm(c) for c in cols}
                fm = {}

                # Match template fields first
                for t in tfields:
                    nt = _norm(t)
                    for c, nc in cols_norm.items():
                        if nc == nt or nc.endswith(nt) or nt in nc:
                            fm[t] = c
                            break

                # If nothing matched, try a set of common target fields
                if not fm:
                    common_targets = ['org_unit_id', 'date', 'metric_key', 'metric_value', 'campaign', 'channel', 'impressions', 'engagements', 'clicks', 'conversions', 'cost', 'event_id', 'name', 'stage', 'count_value']
                    for t in common_targets:
                        nt = _norm(t)
                        for c, nc in cols_norm.items():
                            if nc == nt or nt in nc or nc.endswith(nt) or nc.startswith(nt):
                                fm[t] = c
                                break

                # Fallback: map identical names if no target mapping found
                if not fm:
                    for c in cols:
                        fm[c] = c

                mg = {'field_map': fm}
            else:
                mg = {}

        payload = {'import_job_id': job_id, 'mapping': mg}
        return map_v3(payload)
    conn = db.connect()
    try:
        cur = conn.cursor()
        # store mapping template as a quick save
        cur.execute('INSERT INTO import_mapping_template(name, target_domain, mapping_json, created_by, created_at) VALUES (?,?,?,?,?)', (
            mapping.get('name', f'map_{job_id}'), mapping.get('target_domain','generic'), json.dumps(mapping), mapping.get('created_by'), now_iso()
        ))
        # also store mapping JSON on the import_job for commit
        cur.execute('UPDATE import_job SET status=?, target_domain=?, mapping_json=?, updated_at=? WHERE id=?', ('validating', mapping.get('target_domain','generic'), json.dumps(mapping), now_iso(), job_id))
        conn.commit()
        try:
            audit(conn, mapping.get('created_by') if isinstance(mapping, dict) else None, 'import.map', 'import_job', job_id, {'template': mapping.get('name') if isinstance(mapping, dict) else None, 'target_domain': mapping.get('target_domain') if isinstance(mapping, dict) else None})
        except Exception:
            pass
        return {'import_job_id': job_id, 'status': 'validating'}
    finally:
        conn.close()


@router.post('/import/{job_id}/validate')
def validate_job(job_id: str, options: Dict = {}, allowed_orgs: Optional[list] = Depends(require_scope('STATION'))):
    """Legacy numeric validate endpoint that also accepts v3 import_job_id UUIDs.
    Delegate to v3 `validate_v3` when a UUID is supplied.
    """
    try:
        legacy_id = int(job_id)
    except Exception:
        legacy_id = None
    if legacy_id is None:
        return validate_v3({'import_job_id': job_id})
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT preview_json FROM import_job_preview WHERE import_job_id=? ORDER BY id DESC LIMIT 1', (job_id,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(status_code=400, detail='no preview to validate')
        preview = json.loads(p['preview_json'])
        errors = 0
        warnings = 0
        # simple validation: check for empty rows
        for i, r in enumerate(preview):
            if not any([v for v in r.values()]):
                errors += 1
                cur.execute('INSERT INTO import_job_log(import_job_id, level, message, row_number, created_at) VALUES (?,?,?,?,?)', (job_id, 'error', 'empty row', i+1, now_iso()))
        try:
            cur.execute('UPDATE import_job SET status=?, row_count_detected=?, updated_at=? WHERE id=?', (( 'mapping_required' if errors>0 else 'importing'), len(preview), now_iso(), job_id))
        except Exception:
            try:
                cur.execute('UPDATE import_job SET status=?, updated_at=? WHERE id=?', (( 'mapping_required' if errors>0 else 'importing'), now_iso(), job_id))
            except Exception:
                pass
        conn.commit()
        try:
            audit(conn, None, 'import.validate', 'import_job', job_id, {'errors': errors, 'warnings': warnings})
        except Exception:
            pass
        return {'import_job_id': job_id, 'errors': errors, 'warnings': warnings}
    finally:
        conn.close()


@router.post('/import/{job_id}/commit')
def commit_job(job_id: str, allowed_orgs: Optional[list] = Depends(require_scope('STATION')), current_user: Dict = Depends(get_current_user)):
    """Legacy numeric commit endpoint that also accepts v3 import_job_id UUIDs.
    For UUIDs, delegate to the v3 commit compatibility helper which will
    attempt to locate and commit via the v3 path.
    """
    try:
        legacy_id = int(job_id)
    except Exception:
        legacy_id = None
    if legacy_id is None:
        # call compatibility helper to map v3 id to legacy commit flow
        return commit_v3_compat({'import_job_id': job_id, 'mode': 'append'})
    conn = db.connect()
    try:
        cur = conn.cursor()
        cur.execute('SELECT pj.preview_json, ij.mapping_json, ij.target_domain FROM import_job_preview pj JOIN import_job ij ON pj.import_job_id=ij.id WHERE pj.import_job_id=? ORDER BY pj.id DESC LIMIT 1', (job_id,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(status_code=400, detail='no preview to commit')
        p = db.row_to_dict(cur, p)
        preview = json.loads(p.get('preview_json') or '[]')
        mapping_json = p.get('mapping_json')
        target_domain = p.get('target_domain')

        allowed = set(['funnel_event','cost_benchmark','roi_result','loe','project','projects','task','tasks','meeting','agenda_item','minutes','action_item','decision','calendar_event','event','events','expenses','expense','budget_line_item','fy_budget'])

        row_count = 0
        if mapping_json:
            try:
                mapping = json.loads(mapping_json) if isinstance(mapping_json, str) else mapping_json
            except Exception:
                mapping = mapping_json
        else:
            mapping = None

        target = None
        if mapping and mapping.get('target_table'):
            candidate = mapping.get('target_table')
            # allow explicitly-listed targets OR any real table present in the DB
            try:
                cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (candidate,))
                found = cur.fetchone()
                if found or candidate in allowed:
                    target = candidate
            except Exception:
                # fall back to allowed set only
                if candidate in allowed:
                    target = candidate
        
        field_map = mapping.get('field_map', {})
        if target:
            field_map = mapping.get('field_map', {})
            for r in preview:
                # enforce org scope where possible
                org_val = None
                # check mapped org_unit_id
                if 'org_unit_id' in field_map:
                    org_val = r.get(field_map.get('org_unit_id'))
                # check direct value in row
                if org_val is None:
                    org_val = r.get('org_unit_id') or r.get('org_unit')
                if allowed_orgs is not None and org_val is not None:
                    try:
                        if int(org_val) not in allowed_orgs:
                            raise HTTPException(status_code=403, detail='forbidden')
                    except Exception:
                        # if not an int, compare as string
                        if str(org_val) not in [str(x) for x in allowed_orgs]:
                            raise HTTPException(status_code=403, detail='forbidden')
                # build insert for target
                cols = []
                vals = []
                for tfield, sfield in field_map.items():
                    # translate common alias target fields to actual DB column names
                    mapped_tfield = FIELD_ALIASES.get(tfield, tfield)
                    cols.append(mapped_tfield)
                    vals.append(r.get(sfield))

                # special-case: budget_line_item may reference fy/org_unit; create/find fy_budget when mapping provides fy/org_unit
                if target == 'budget_line_item':
                    # if mapping provides fy and org_unit_id fields, try to find/create fy_budget and set fy_budget_id
                    fy_val = None
                    org_val = None
                    if 'fy' in field_map:
                        fy_val = r.get(field_map.get('fy'))
                    if 'org_unit_id' in field_map:
                        org_val = r.get(field_map.get('org_unit_id'))
                    try:
                        if fy_val is not None and org_val is not None:
                            cur.execute('SELECT id FROM fy_budget WHERE org_unit_id=? AND fy=?', (int(org_val), int(fy_val)))
                            frow = cur.fetchone()
                            if frow:
                                fy_budget_id = frow['id'] if 'id' in frow.keys() else frow[0]
                            else:
                                cur.execute('INSERT INTO fy_budget(org_unit_id,fy,total_allocated,created_at) VALUES (?,?,?,?)', (int(org_val), int(fy_val), 0.0, now_iso()))
                                fy_budget_id = cur.lastrowid
                            # ensure fy_budget_id inserted as first column if not present
                            cols.insert(0, 'fy_budget_id')
                            vals.insert(0, fy_budget_id)
                    except Exception:
                        pass

                # add provenance fields if present in table
                cols.extend(['created_at','updated_at','import_job_id'])
                vals.extend([now_iso(), now_iso(), job_id])
                placeholders = ','.join(['?'] * len(vals))
                sql = f"INSERT OR REPLACE INTO {target}({','.join(cols)}) VALUES ({placeholders})"
                try:
                    cur.execute(sql, tuple(vals))
                    row_count += 1
                except Exception:
                    # try a permissive insert: only insert fields that actually exist in the table
                    try:
                        cur.execute(f"PRAGMA table_info({target})")
                        existing_cols = [c[1] for c in cur.fetchall()]
                        insert_cols = [c for c in cols if c in existing_cols]
                        insert_vals = [vals[i] for i, c in enumerate(cols) if c in existing_cols]
                        if insert_cols:
                            placeholders = ','.join(['?'] * len(insert_vals))
                            sql = f"INSERT OR REPLACE INTO {target}({','.join(insert_cols)}) VALUES ({placeholders})"
                            cur.execute(sql, tuple(insert_vals))
                            row_count += 1
                    except Exception:
                        pass
            # legacy schema may not have `row_count_imported`; use `row_count_detected` as a compatible column
            try:
                cur.execute('UPDATE import_job SET row_count_imported=?, status=?, updated_at=? WHERE id=?', (row_count, 'completed', now_iso(), job_id))
            except Exception:
                try:
                    cur.execute('UPDATE import_job SET row_count_detected=?, status=?, updated_at=? WHERE id=?', (row_count, 'completed', now_iso(), job_id))
                except Exception:
                    pass
            conn.commit()
            try:
                audit(conn, None, 'import.commit', 'import_job', job_id, {'imported': row_count, 'target_table': target})
            except Exception:
                pass
            # mark any provenance rows for this job as processed
            try:
                cur.execute("PRAGMA table_info(imported_rows)")
                cols = [c[1] for c in cur.fetchall()]
                if 'processed' in cols:
                    # mark as processed and record actor where available
                    try:
                        uname = current_user.get('username') if isinstance(current_user, dict) else None
                        if 'processed_by' in cols:
                            cur.execute('UPDATE imported_rows SET processed=1, processed_at=?, processed_by=? WHERE import_job_id=?', (now_iso(), uname or 'system', job_id))
                        else:
                            cur.execute('UPDATE imported_rows SET processed=1, processed_at=? WHERE import_job_id=?', (now_iso(), job_id))
                    except Exception:
                        # best-effort; continue without failing commit
                        pass
                    conn.commit()
            except Exception:
                pass
            return {'import_job_id': job_id, 'imported': row_count, 'target_table': target}

        # fallback: store into imported_rows for provenance
        for r in preview:
            cur.execute('INSERT INTO imported_rows(import_job_id, target_domain, row_json, created_at) VALUES (?,?,?,?)', (job_id, target_domain or 'generic', json.dumps(r), now_iso()))
            row_count += 1
        try:
            cur.execute('UPDATE import_job SET row_count_imported=?, status=?, updated_at=? WHERE id=?', (row_count, 'completed', now_iso(), job_id))
        except Exception:
            try:
                cur.execute('UPDATE import_job SET row_count_detected=?, status=?, updated_at=? WHERE id=?', (row_count, 'completed', now_iso(), job_id))
            except Exception:
                pass
        conn.commit()
        try:
            audit(conn, None, 'import.commit', 'import_job', job_id, {'imported': row_count, 'fallback': True})
        except Exception:
            pass
        return {'import_job_id': job_id, 'imported': row_count}
    finally:
        conn.close()


@router.post('/import/compat/commit_v3')
def commit_v3_compat(payload: Dict[str, Any] = Body(...)):
    """Compatibility helper: given a v3 `import_job_id` attempt to locate the
    legacy numeric import_job and invoke the legacy commit logic so callers
    using v3 ids can complete a commit immediately."""
    import_job_id = payload.get('import_job_id') if payload else None
    mode = payload.get('mode', 'append') if payload else 'append'
    if not import_job_id:
        raise HTTPException(status_code=400, detail='missing import_job_id')
    conn = db.connect()
    try:
        cur = conn.cursor()
        # try to find a matching legacy import_job by filename or sha
        cur.execute('SELECT filename, file_sha256 FROM import_job_v3 WHERE id=? LIMIT 1', (import_job_id,))
        j = cur.fetchone()
        # normalize DB row into a plain dict to avoid sqlite3.Row quirks
        try:
            j = db.row_to_dict(cur, j) if j is not None else None
        except Exception:
            # fallback: attempt to build a dict from cursor description
            try:
                if j is not None:
                    desc = cur.description
                    j = {desc[i][0]: j[i] for i in range(len(desc))}
                else:
                    j = None
            except Exception:
                j = None
        legacy_id = None
        if j:
            # j should now be a plain dict; use safe get() calls
            filename = j.get('filename') or j.get('filename_original')
            sha = j.get('file_sha256') or j.get('sha256') or j.get('file_hash')
            if filename:
                cur.execute('SELECT id FROM import_job WHERE filename_original=? OR filename=? LIMIT 1', (filename, filename))
                f = cur.fetchone()
                try:
                    f = db.row_to_dict(cur, f) if f is not None else None
                except Exception:
                    try:
                        if f is not None:
                            desc = cur.description
                            f = {desc[i][0]: f[i] for i in range(len(desc))}
                        else:
                            f = None
                    except Exception:
                        f = None
                if f:
                    legacy_id = f.get('id')
            if not legacy_id and sha:
                cur.execute('SELECT id FROM import_job WHERE sha256_hash=? OR file_hash=? LIMIT 1', (sha, sha))
                f = cur.fetchone()
                try:
                    f = db.row_to_dict(cur, f) if f is not None else None
                except Exception:
                    try:
                        if f is not None:
                            desc = cur.description
                            f = {desc[i][0]: f[i] for i in range(len(desc))}
                        else:
                            f = None
                    except Exception:
                        f = None
                if f:
                    legacy_id = f.get('id')
        if not legacy_id:
            raise HTTPException(status_code=404, detail='legacy job not found for import_job_id')
        # Copy any legacy `imported_rows` into the v3 import_job_id so the v3
        # commit path can consume them. This avoids needing a preview when the
        # legacy job still holds imported_rows and matches by filename/sha.
        try:
            cur.execute('SELECT row_json, target_domain, created_at FROM imported_rows WHERE import_job_id=?', (legacy_id,))
            legacy_rows = cur.fetchall()
            try:
                print(f"commit_v3_compat: found {len(legacy_rows)} legacy_rows for legacy_id={legacy_id}")
            except Exception:
                pass
            for lr in legacy_rows:
                try:
                    lrd = db.row_to_dict(cur, lr) if lr else None
                except Exception:
                    try:
                        if lr is not None:
                            desc = cur.description
                            lrd = {desc[i][0]: lr[i] for i in range(len(desc))}
                        else:
                            lrd = None
                    except Exception:
                        lrd = None
                row_json = lrd.get('row_json') if lrd else None
                # If legacy row didn't expose a 'row_json' value, fall back
                # to JSON-serializing the available row dict so the v3
                # commit path can parse it into a mapping-like object.
                if row_json is None and lrd:
                    try:
                        row_json = json.dumps(lrd)
                    except Exception:
                        row_json = None
                target_domain = lrd.get('target_domain') if lrd else None
                created_at = lrd.get('created_at') if lrd else None
                try:
                    cur.execute('INSERT INTO imported_rows(import_job_id, target_domain, row_json, created_at) VALUES (?,?,?,?)', (import_job_id, target_domain or 'production', row_json, created_at or now_iso()))
                except Exception:
                    pass
            conn.commit()
            try:
                cur.execute('SELECT COUNT(*) FROM imported_rows WHERE import_job_id=?', (import_job_id,))
                cnt = cur.fetchone()
                try:
                    print(f"commit_v3_compat: inserted {cnt[0] if cnt else 'unknown'} rows for import_job_id={import_job_id}")
                except Exception:
                    pass
            except Exception:
                pass
            # Now call v3 commit logic which will consume imported_rows for the v3 id
            # Provide a minimal `current_user` dict when invoking the endpoint
            # function directly so code paths that call `.get()` work when the
            # FastAPI dependency injection context is not present.
            res = commit_v3({'import_job_id': import_job_id, 'mode': mode}, None, {'username': 'system'})
            return {'status': 'ok', 'import_job_id': import_job_id, 'legacy_job_id': legacy_id, 'commit_result': res}
        except HTTPException:
            raise
        except Exception as e:
            # Log stack for debugging in test runs to trace unexpected errors
            try:
                import traceback
                print('commit_v3_compat: internal exception:\n' + traceback.format_exc())
            except Exception:
                pass
            raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
